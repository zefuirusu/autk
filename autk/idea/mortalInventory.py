#!/usr/bin/env python
# coding=utf-8
'''
这是2021年9月1日从别处复制过来的，上次编辑是2021年7月4日下午12:51:32；
Based on Perpetual System for Inventory Booking.
When using this script, you need to overwrite InvRow.classify() and Inventory.start().
'''
import os
import threading
import datetime
from json import load
from decimal import Decimal
from numpy import nan,array
from pandas import read_excel,concat,Series,DataFrame
from autk.parser.funcs import get_time_str
def mlogger(afunc,log_path='./inventory_log.txt'):
    import datetime
    # log_start_time=datetime.datetime.now()
    # log_start_time=str(log_start_time)
    # with open(log_path,mode='w',encoding='utf-8') as f:
    #     f.write('log start at:\t')
    #     f.write(log_start_time)
    #     f.write('\n')
    def one_write(w_str):
        with open(log_path,mode='a',encoding='utf-8') as f:
            f.write(w_str)
            f.write('\n')
        pass
    def log_wrapper(*args,**kwargs):
        t1=datetime.datetime.now()
        # one_write(t1)
        resu=afunc(*args,**kwargs)
        t2=datetime.datetime.now()
        # one_write(t2)
        t_interval=t2-t1
        t_interval=str(t_interval)
        w_str=r'[function]'+r'['+afunc.__name__+r']:'+'time spent\t'+t_interval
        one_write(w_str)
        return resu
    return log_wrapper
class MortalInventory:
    def __init__(self,xlmeta=None,item_json_path='./inventory_item.json'):
        '''
        parameters:
            xlmeta:meta data of Excel Books, can be a dict, a list (single file), or multi-dimension list with 3 columns;
            xlmeta like:
                {
                    file_path_1:[
                        [sheet_name_1,0],
                        [sheet_name_2,4],
                        ...
                    ],
                    file_path_2:[
                        [sheet_name_2,3],
                        [],
                        ...
                    ],
                    file_path_3:True,
                    ...
                }
            item_json_path:to indicate columns used in calculation.
        '''
        # Table structure JSON.
        with open(item_json_path,mode='r',encoding='utf-8') as f:
            item_json=load(f)
        self.item_json=item_json
        self.item_keys=list(self.item_json.keys())
        self.item_values=list(self.item_json.values())
        # Initializing with [file_path,sheet_name,header] attributes:
        self.xlmeta=xlmeta
        self.pathli=[]
        print('='*5)
        print('Initializing meta data:',self.xlmeta)
        if isinstance(self.xlmeta,dict):
            for file_path in self.xlmeta:
                for sheet_meta in self.xlmeta[file_path]:
                    shtna=sheet_meta[0]
                    title=sheet_meta[1]
                    self.pathli.append([file_path,shtna,title])
            print('Initialized with json:',self.pathli)
        elif isinstance(self.xlmeta,list):
            xlmeta_fake=array(self.xlmeta)
            if xlmeta_fake.ndim==1:
                self.pathli.append(self.xlmeta)
            elif xlmeta_fake.ndim==2:
                self.pathli=self.xlmeta
            else:
                pass
            print('Initialized with a single file:',self.pathli)
        else:
            pass
        print('-'*5)
        self.raw_data=None # whole data in the form of pandas.DataFrame.
        self.load_count=0
        # Core-data attributes:
        self.data=None # updating at anytime.
        self.columns=None
        self.price_matrix=None
        self.data_set=[]# data prepared for multi-thread-data-processing, in which elements are pandas.DataFrame.
        self.row_set=[] # data cache for each thread of row-iteration, in which elements are class 'InvRow'.
        # Detailed-data attributes:
        self.years=None
        self.months=None
        self.material_items=None # raw materials
        self.product_items=None # product project
        self.material_json={}
        self.product_json={}
        # Temporary data attributes:
        self.row_temp=[] # data cache for each thread of filter_sum, after which self.row_temp will be cleared by self.__clear_row_temp(). Elements are pandas.Series yielded by DataFrame.iterrows().
        self.data_temp=[] # used in self.cal_price.
        self.meta_cols=[] # seems useless, used in self.get_pvt.
        self.data_cols=[] # seems useless, used in self.get_pvt.
        pass
    def __time_monitor(self,afunc):
        import datetime
        def time_wrapper(*args,**kwargs):
            t1=datetime.datetime.now()
            print('[',t1,']','start_method:',afunc.__name__)
            afunc(*args,**kwargs)
            t2=datetime.datetime.now()
            print('[',t2,']','end_method:',afunc.__name__)
            return afunc(*args,**kwargs)
        return time_wrapper
    @__time_monitor
    def load_df(self,df):
        self.data=df
        self.data_set=[df]
        self.load_count=1
        pass
    @staticmethod
    def __set_date_to_str(row_series):
        date=str(row_series['date'])
        return date
    @staticmethod
    def __set_month(row_series):
        from numpy import nan
        date_str=row_series['date']
        if date_str is not nan:
            month=date_str[5:7]
        else:
            month=nan
        return month
    @staticmethod
    def __set_year(row_series):
        from numpy import nan
        date_str=row_series['date']
        if date_str is not nan:
            year=date_str[0:4]
        else:
            year=nan
        return year
    @staticmethod
    def __trans_nan(num):
        '''
        Transform nan into 0.0;
        '''
        from numpy import nan
        if num is nan:
            return 0.0
        else:
            return num
    @staticmethod
    def __set_start_num(row_series):
        end_num=MortalInventory.__trans_nan(row_series['end_num'])
        out_num=MortalInventory.__trans_nan(row_series['out_num'])
        in_num=MortalInventory.__trans_nan(row_series['in_num'])
        start_num=end_num+out_num-in_num
        return start_num
    @staticmethod
    def __set_start_amount(row_series):
        end_num=MortalInventory.__trans_nan(row_series['end_amount'])
        out_num=MortalInventory.__trans_nan(row_series['out_amount'])
        in_num=MortalInventory.__trans_nan(row_series['in_amount'])
        start_amount=end_num+out_num-in_num
        return start_amount
    @staticmethod
    def __set_start_price(row_series):
        start_num=MortalInventory.__trans_nan(row_series['start_num'])
        # start_num=Decimal(str(start_num))
        # start_num=float(start_num)
        start_amount=MortalInventory.__trans_nan(row_series['start_amount'])
        # start_amount=Decimal(str(start_amount))
        # start_amount=float(start_amount)
        if isinstance(start_num,str):
            print('**',row_series)
        if isinstance(start_amount,str):
            print('**',row_series)
        if start_num !=0:
            return start_amount/start_num
        else:
            return 0.0
    def __load_data_once(self,fpath,shtna,title):
        d=read_excel(fpath,sheet_name=shtna,header=title,engine='openpyxl')
        d_fake=DataFrame([],columns=list(self.item_json.keys()))
        for col in d_fake.columns:
            if self.item_json[col] is not None:
                d_fake[col]=d[d.columns[self.item_json[col]]]
            else:
                pass
        d=d_fake
        if self.item_json['date'] is not None:
            d['date']=d.apply(MortalInventory.__set_date_to_str,axis=1)
            d['date']=d['date'].fillna(method='ffill')
        else:
            pass
        if self.item_json['month'] is None:
            d['month']=d.apply(MortalInventory.__set_month,axis=1)
            d['month']=d['month'].fillna(method='ffill')
        else:
            pass
        if self.item_json['year'] is None:
            d['year']=d.apply(MortalInventory.__set_year,axis=1)
            d['year']=d['year'].fillna(method='ffill')
        else:
            pass
        if self.item_json['start_num'] is None:
            d['start_num']=d.apply(MortalInventory.__set_start_num,axis=1)
        else:
            pass
        if self.item_json['start_amount'] is None:
            d['start_amount']=d.apply(MortalInventory.__set_start_amount,axis=1)
        else:
            pass
        if self.item_json['start_price'] is None:
            d['start_price']=d.apply(MortalInventory.__set_start_price,axis=1)
        else:
            pass
        self.data_set.append(d)
        pass
    def clear_core_data(self):
        self.raw_data=None # whole data in the form of pandas.DataFrame.
        self.load_count=0
        # Core-data attributes:
        self.data=None # updating at anytime.
        self.columns=None
        self.price_matrix=None
        self.data_set=[]# data prepared for multi-thread-data-processing, in which elements are pandas.DataFrame.
        self.row_set=[] # data cache for each thread of row-iteration, in which elements are class 'InvRow'.
        # Detailed-data attributes:
        self.years=None
        self.months=None
        self.material_items=None # raw materials
        self.product_items=None # product project
        self.material_json={}
        self.product_json={}
        # Temporary data attributes:
        self.row_temp=[] # data cache for each thread of filter_sum, after which self.row_temp will be cleared by self.__clear_row_temp(). Elements are pandas.Series yielded by DataFrame.iterrows().
        self.data_temp=[] # used in self.cal_price.
        self.meta_cols=[] # seems useless, used in self.get_pvt.
        self.data_cols=[] # seems useless, used in self.get_pvt.
        pass
    @__time_monitor
    def load_raw_data(self):
        if self.load_count >0:
            self.clear_core_data()
        else:
            pass # load data first time.
        thread_list=[]
        for i in range(len(self.pathli)):
            fpath=self.pathli[i][0]
            shtna=self.pathli[i][1]
            title=self.pathli[i][2]
            thread_name=r'/'.join([fpath.split(os.sep)[-1],shtna,str(title)])
            th=threading.Thread(target=self.__load_data_once,args=(fpath,shtna,title),name=thread_name)
            thread_list.append(th)
            continue
        for t in thread_list:
            t.start()
            print('reading: %s'%t.name)
            continue
        for t in thread_list:
            t.join()
            continue
        self.raw_data=concat(self.data_set,axis=0,join='outer')
        self.data=self.raw_data
        self.raw_data=None
        self.columns=list(self.data.columns)
        def get_pure_list(col):
            from numpy import nan
            a_list=list(self.data[col].drop_duplicates())
            if nan in a_list:
                a_list.remove(nan)
            return a_list
        self.years=get_pure_list('year')
        self.months=get_pure_list('month')
        self.material_items=get_pure_list('material')
        self.product_items=get_pure_list('product')
        for y in self.years:
            self.material_json[y]={}
            for mater in self.material_items:
                self.material_json[y][mater]={}
                for month in self.months:
                    self.material_json[y][mater][month]=DataFrame([],columns=self.columns)
        self.load_count+=1
        pass
    def __load_dir_data_once(self,fpath,shtna,title):
        d=read_excel(fpath,sheet_name=shtna,header=title,engine='openpyxl')
        # print(d.columns)
        d_fake=DataFrame([],columns=list(self.item_json.keys()))
        for col in d_fake.columns:
            if self.item_json[col] is not None:
                # print(d.iloc[:,self.item_json[col]])
                d_fake[col]=d[d.columns[self.item_json[col]]]
                # d_fake[col]=d.iloc[:,self.item_json[col]]
            else:
                pass
        d=d_fake
        d['month']=d.apply(MortalInventory.__set_month,axis=1)
        d['month']=d['month'].fillna(method='ffill')
        d['year']=d.apply(MortalInventory.__set_year,axis=1)
        d['year']=d['year'].fillna(method='ffill')
        d['start_num']=d.apply(MortalInventory.__set_start_num,axis=1)
        d['start_amount']=d.apply(MortalInventory.__set_start_amount,axis=1)
        d['start_price']=d.apply(MortalInventory.__set_start_price,axis=1)
        self.data_set.append(d)
        pass
    @__time_monitor
    def prescan(self,load_dir,title=0):
        from openpyxl import load_workbook
        pathli=[]
        for f in os.listdir(load_dir):
            pathli.append(os.path.join(load_dir,f))
            continue
        print('reading files:','\n',pathli)
        print('-'*5)
        for f in pathli:
            print(f.split(os.sep)[-1])
            shtli=load_workbook(f).sheetnames
            for sht in shtli:
                print('\t',sht)
                print('\t'*2,read_excel(f,sht,header=title).columns)
        print('-'*5)
        pass
    @__time_monitor
    def load_dir_data(self,load_dir,title=0):
        '''
        Load all sheets of all files in load_dir;
        '''
        from openpyxl import load_workbook
        self.clear_core_data()
        pathli=[]
        for f in os.listdir(load_dir):
            pathli.append(os.path.join(load_dir,f))
            continue
        print('-'*5)
        print('reading files:','\n',pathli)
        thread_list=[]
        for f in pathli:
            shtli=load_workbook(f).sheetnames
            for sht in shtli:
                t=threading.Thread(target=self.__load_dir_data_once,args=(f,sht,title))
                thread_list.append(t)
        print('-'*5)
        for t in thread_list:
            t.start()
        for t in thread_list:
            t.join()
        self.raw_data=concat(self.data_set,axis=0,join='outer')
        self.data=self.raw_data
        self.raw_data=None
        self.columns=list(self.data.columns)
        def get_pure_list(col):
            from numpy import nan
            a_list=list(self.data[col].drop_duplicates())
            if nan in a_list:
                a_list.remove(nan)
            return a_list
        self.years=get_pure_list('year')
        self.months=get_pure_list('month')
        self.material_items=get_pure_list('material')
        self.product_items=get_pure_list('product')
        for mater in self.material_items:
            self.material_json[mater]={}
            for y in self.years:
                self.material_json[mater][y]={}
                for month in self.months:
                    self.material_json[mater][y][month]=DataFrame([],columns=self.columns)
        self.load_count+=1
        pass
    def __cluster_material_once(self,df):
        for i in df.iterrows():
            row_data=i[1]
            mater=row_data['material']
            year=row_data['year']
            month=row_data['month']
            row_data_fake=DataFrame(row_data).T
            row_data.columns=row_data.index
            if year in self.years and month in self.months and mater in self.material_items:
                current_df=self.material_json[year][mater][month]
                self.material_json[year][mater][month]=current_df.append(row_data_fake)
                # self.material_json[mater][year][month]=concat([current_df,row_data_fake],axis=0,join='outer')
            else:
                pass
        pass
    def cluster_material_once(self,df):
        self.__cluster_material_once(df)
        return
    @__time_monitor
    def cluster_material(self):
        if len(self.data_set)==0:
            print('[Warning:]','nothing in self.data_set!')
            return
        else:
            pass
        thread_list=[]
        for df in self.data_set:
            t=threading.Thread(target=self.__cluster_material_once,args=(df,))
            thread_list.append(t)
            continue
        for t in thread_list:
            t.start()
            print('start clustering...',t)
        for t in thread_list:
            t.join()
        return
    def __allocated_month(self):
        '''
        # what if '>本月合计' not included in columns?
        '''
        from numpy import nan
        self.__clear_row_temp()
        if self.data is None:
            print('[Warning:]','Load raw data first!')
        else:
            pass
        pass
    def get_material_data(self,year,material,month):
        '''
        self.material_json must not be empty, so self.cluster_material() before calling this method.
        '''
        return self.material_json[year][material][month]
    def __get_month_price_once(self,month_df):
        '''
            1.Filter data and get pure transaction records;
            2.Calculate start_num at the beginning of the month;
            3.Get start_amount from last month's DataFrame;
            4.Calculate 'recal_end_price', 'recal_end_amount', and 'recal_amount_allocated' for current month_df.
        '''
        self.price_json={}
        pass
    def get_month_price(self):
        '''
        Call self.cluster_material() and then, calculate average price for each material_item in every month.
        Monthly average price data will be updated into self.price_json.
        '''
        thread_list=[]
        for y in self.years:
            for mater in self.material_items:
                for m in self.months:
                    t=threading.Thread(target=self.__get_month_price_once,args=(m,))
                    thread_list.append(t)
        for t in thread_list:
            t.start()
        for t in thread_list:
            t.join()
        pass
    def __cal_price_once(self,df,moving_weighted_average=True,over_write=False):
        '''
        Each DataFrame of self.data_set calculates independently.
        '''
        from numpy import nan
        self.__clear_row_temp()
        if self.data is None:
            print('[Warning:]','Load raw data first!')
        else:
            pass
        d_start=self.__filter_once(df,[[r'^.*期初.*$','summary',True,True]])
        if d_start.shape[0] !=1:
            print('[Warning:]','start_amount is not unique!')
        else:
            pass
        if moving_weighted_average==True:
            # print('Moving Weighted Average Method.')
            d_body=self.__filter_once(df,[[r'^[^n].*','date',True,True]])
        else:
            # print('Weighted Average by Month.')
            # what if '>本月合计' not included in columns?
            if '>本月合计' in list(df['summary'].drop_duplicates()):
                d_body=self.__filter_once(df,[['>本月合计','summary',False,False]])
            else:
                # __allocated_month
                def get_to_sum(col_str):
                    a=list(df[col_str].fillna('woc').drop_duplicates())
                    if 'woc' in a:
                        a.remove('woc')
                    else:
                        pass
                    return a
                months=get_to_sum('month')
                years=get_to_sum('year')
                materials=get_to_sum('material')
                def yield_sumed_df():
                    for y in years:
                        for m in months:
                            for mater in materials:
                                d_m=self.__filter_once(df,[[y,'year',False,False],[m,'month',False,False],[mater,'material',False,False]])
                                # d_small_body=DataFrame([],columns=d_start.columns)
                                d_small_body=DataFrame([],columns=d_start.columns)
                                d_small_body.loc[0,'year']=y
                                d_small_body.loc[0,'month']=m
                                d_small_body.loc[0,'material']=mater
                                from numpy import float64
                                for col in d_small_body.columns:
                                    if d_small_body[col].dtype==float64:
                                        d_small_body.loc[0,col]=d_m[col].sum(axis=0)
                                    else:
                                        pass
                                yield d_small_body
                d_body=concat(yield_sumed_df(),axis=0,join='outer')
                pass
        self.__clear_row_temp()
        d=concat([d_start,d_body],axis=0,join='outer')
        start_bal=d.loc[d.index[0],'end_amount']
        d.loc[d.index[0],'recal_end_amount']=start_bal
        for i in range(1,d.shape[0]):
            amount_temp=d.loc[d.index[i-1],'recal_end_amount']+d.loc[d.index[i],'in_amount']
            num_temp=d.loc[d.index[i-1],'end_num']+d.loc[d.index[i],'in_num']
            if num_temp !=0:
                price_temp=amount_temp/num_temp
            else:
                price_temp=nan
            d.loc[d.index[i],'recal_end_price']=price_temp
            d.loc[d.index[i],'recal_end_amount']=price_temp*d.loc[d.index[i],'end_num']
            continue
        self.data_temp.append(d)
        if over_write==True:
            df=d
        else:
            pass
    def __cal_price_once_decimal(self,df,moving_weighted_average=True,over_write=False):
        '''
        Accurate Mode on when calculating price.
        '''
        print('accurate mode: cal_price.')
        from numpy import nan
        self.__clear_row_temp()
        if self.data is None:
            print('[Warning:]','Load raw data first!')
        else:
            pass
        d_start=self.__filter_once(df,[[r'^.*期初.*$','summary',True,True]])
        if moving_weighted_average==True:
            # print('Moving Weighted Average Method.')
            d_body=self.__filter_once(df,[[r'^[^n].*','date',True,True]])
        else:
            # print('Weighted Average by Month.')
            d_body=self.__filter_once(df,[['>本月合计','summary',False,False]])
        self.__clear_row_temp()
        d=concat([d_start,d_body],axis=0,join='outer')
        def trans_decimal(num):
            return Decimal(str(num))
        start_bal=trans_decimal(d.loc[d.index[0],'end_amount'])
        d.loc[d.index[0],'recal_end_amount']=start_bal
        for i in range(1,d.shape[0]):
            amount_temp=trans_decimal(d.loc[d.index[i-1],'end_amount'])+trans_decimal(d.loc[d.index[i],'in_amount'])
            num_temp=trans_decimal(d.loc[d.index[i-1],'end_num'])+trans_decimal(d.loc[d.index[i],'in_num'])
            if num_temp !=0:
                price_temp=amount_temp/num_temp
            else:
                price_temp=nan
            d.loc[d.index[i],'recal_end_price']=price_temp
            d.loc[d.index[i],'recal_end_amount']=trans_decimal(price_temp)*trans_decimal(d.loc[d.index[i],'end_num'])
            continue
        self.data_temp.append(d)
        if over_write==True:
            df=d
        else:
            pass
    # @__time_monitor
    def cal_price(self,moving_weighted_average=True,over_write=False,accurate=False):
        '''
        This method will not update self.price_matrix;
        Moving Weighted Average is set to default.
        return:
            DataFrame of self.data with column 'recal_end_price' and 'recal_end_amount'.
        '''
        self.data_temp=[]
        thread_list=[]
        for df in self.data_set:
            if accurate==False:
                th=threading.Thread(target=self.__cal_price_once,args=(df,moving_weighted_average,True))
            else:
                th=threading.Thread(target=self.__cal_price_once_decimal,args=(df,moving_weighted_average,True))
            thread_list.append(th)
            continue
        for t in thread_list:
            t.start()
        for t in thread_list:
            t.join()
        if len(self.data_temp)==0:
            d=DataFrame([],columns=self.data_set[0].columns)
        else:
            d=concat(self.data_temp,axis=0,join='outer')
        self.data_temp=[]
        if over_write==True:
            self.data=d
        else:
            pass
        return d
    # @__time_monitor
    def get_price_matrix(self,moving_weighted_average=False):
        '''
        This method will over_write self.price_matrix, which structure differs when choosing 'moving_weighted_average' or not.
        '''
        price_matrix=self.cal_price(moving_weighted_average=moving_weighted_average,over_write=False,accurate=False)
        if moving_weighted_average==True:
            price_key_cols=['material_id','material','date','recal_end_price']
            price_matrix=price_matrix[price_key_cols]
            price_matrix=price_matrix.fillna(0.0)
            price_matrix=price_matrix[price_matrix['date']!=0]
        else:
            price_key_cols=['material_id','material','year','month','recal_end_price']
            price_matrix=price_matrix[price_key_cols]
            price_matrix=price_matrix.fillna(0.0)
            price_matrix=price_matrix[price_matrix['month']!=0]
        self.price_matrix=price_matrix
        return price_matrix
    # @staticmethod
    def __set_price_month(self,row_series):
        p=self.get_price_matrix(moving_weighted_average=False)
        y=row_series['year']
        m=row_series['month']
        mater=row_series['material']
        end_price_row_set=[]
        for i in p.iterrows():
            if_con=[]
            row_data=i[1]
            if_con.append(row_data['year']==y)
            if_con.append(row_data['month']==m)
            if_con.append(row_data['material']==mater)
            if if_con==[True,True,True]:
                end_price_row_set.append(row_data)
                price_month=row_data['recal_end_price']
                row_series['recal_end_price']=price_month
                return price_month
    @__time_monitor
    def cal_amount_allocated(self,moving_weighted_average=False):
        '''
        Calculate amount allocated from credit of material cost and write into column 'recal_amount_allocated'.
        This method will overwrite self.data.
        parameters:
            moving_weighted_average: True or False;
        return:
            Updated self.data.
        '''
        if moving_weighted_average==True:
            p=self.get_price_matrix(moving_weighted_average=moving_weighted_average)
            self.cal_price(moving_weighted_average=moving_weighted_average,over_write=True)
            self.data['recal_amount_allocated']=self.data['recal_end_price']*self.data['out_num']
            return self.data
        else:
            # Monthly average:
            p=self.get_price_matrix(moving_weighted_average=moving_weighted_average)
            print('check unique of price_matrix:',p.shape[0]==p.drop_duplicates().shape[0])
            self.data['recal_end_price']=self.data.apply(self.__set_price_month,axis=1)
            self.data['recal_end_amount']=self.data['recal_end_price']*self.data['end_num']
            self.data['recal_amount_allocated']=self.data['recal_end_price']*self.data['out_num']
            return self.data
    def __filter_once(self,df,condition_matrix):
        '''
        parameters:
            condition_matrix, a matrix with several rows but only 4 columns; like [[condition,label,regex,match],[condition,label,regex,match],...];
        '''
        once_result=[]
        for i in df.iterrows():
            row_data=i[1]
            condition_set=[]
            for con in condition_matrix:
                regitem=con[0]
                compare_item=con[1]
                regex=con[2]
                match=con[3]
                if regex==False:
                    if regitem==row_data[compare_item]:
                        condition_set.append(True)
                    else:
                        condition_set.append(False)
                else:
                    import re
                    regitem=re.compile(regitem)
                    if match==True:
                        b=re.match(regitem,str(row_data[compare_item]))
                    else:
                        b=re.search(regitem,str(row_data[compare_item]))
                    if b is not None:
                        condition_set.append(True)
                    else:
                        condition_set.append(False)
            if condition_set==[True]*len(condition_set):
                self.row_temp.append(row_data)
                once_result.append(row_data)
            else:
                pass
            continue
        resu=concat(once_result,axis=1).T
        return resu
    def __clear_row_temp(self):
        self.row_temp=[]
        return
    @__time_monitor
    def filter(self,condition_matrix,over_write=False):
        '''
        parameters:
            self.data_set must not be None.
            condition_matrix, a matrix with several rows but only 2 columns; like [[condition,label],[condition,label]];
            regular expression is supported, with match mode set to default;
        return:
            DataFrame of filtered data by condition 'condition_matrix';
        '''
        thread_list=[]
        self.__clear_row_temp()
        for df in self.data_set:
            th=threading.Thread(target=self.__filter_once,args=(df,condition_matrix))
            thread_list.append(th)
            continue
        for t in thread_list:
            t.start()
        for t in thread_list:
            t.join()
        if len(self.row_temp)==0:
            d=DataFrame([],columns=self.data.columns)
        else:
            d=concat(self.row_temp,axis=1).T
        self.__clear_row_temp()
        if over_write==True:
            self.data=d
        else:
            pass
        return d
    @__time_monitor
    def filter_data(self,condition_matrix,over_write=False):
        '''
        Filtering according to self.data.
        '''
        if self.data is None:
            self.load_raw_data()
        else:
            pass
        df=self.data
        if over_write==True:
            resu=self.__filter_once(df,condition_matrix)
            self.load_df(resu)
            return resu
        else:
            return self.__filter_once(df,condition_matrix)
    @__time_monitor
    def recal_product_cost(self):
        '''
        Re-calculate cost of product and return a 'sum_json' like: {product:sum_of_allocated_amount};
        '''
        from numpy import nan
        from copy import deepcopy
        m=deepcopy(self)
        m.filter_data([[r'^\d.*$','date',True,True],[r'^[^n].*$','product',True,True]],over_write=True)
        prod_list=list(m.data['product'].drop_duplicates())
        if nan in prod_list:
            prod_list.remove(nan)
        else:
            pass
        sum_json={}
        for product in prod_list:
            product_recal_amount_allocated=m.filter_data([[product,'product',False,False]])['recal_amount_allocated'].sum(axis=0)
            sum_json[product]=product_recal_amount_allocated
            # sum_json[product]=product_json[product]['recal_amount_allocated'].sum(axis=0)
            print('sum for %s'%product,product_recal_amount_allocated)
        return sum_json
    @__time_monitor
    def export_material_data(self,savepath):
        '''
        Save self.data, self.price_matrix, product_table and material_table to a Microsoft Excel file.
        '''
        from pandas import ExcelWriter
        from openpyxl import Workbook,load_workbook
        if os.path.isfile(savepath)==True:
            wb=load_workbook(savepath)
        elif os.path.isdir(savepath)==True:
            file_name=r'-'.join(['inventory',get_time_str(),r'.xlsx'])
            savepath=os.path.join(file_name,savepath)
            wb=Workbook()
            wb.save(savepath)
        else:
            wb=Workbook()
            wb.save(savepath)
        wter=ExcelWriter(savepath,engine='openpyxl')
        wter.book=wb
        if self.item_json['date'] is not None:
            d=self.filter_data([[r'^\d.+$','date',True,True]],over_write=False)
        else:
            d=self.data
        d_material=d.set_index(['year','material','month']).sort_values(['year','material','month'],ascending=True,axis=0)
        d_product=d.set_index(['year','product','month']).sort_values(['year','product','month'],ascending=True,axis=0)
        from numpy import sum
        d_pivot_product=d_product.pivot_table(values=['recal_amount_allocated'],index=['year','material'],columns=['product'],aggfunc=sum)
        print(d_pivot_product)
        self.data.to_excel(wter,sheet_name='inventory')
        # self.price_matrix.to_excel(wter,sheet_name='price')
        d_material.to_excel(wter,sheet_name='material')
        d_product.to_excel(wter,sheet_name='product')
        d_pivot_product.to_excel(wter,sheet_name='material-product-pvt')
        # for mater in self.material_json:
        #     for y in self.material_json[mater]:
        #         for m in self.material_json[mater][y]:
        #             month_df=m
        wter.save()
        return
    @__time_monitor
    def export_json_material(self,savepath,depth='year'):
        '''
        Export self.material_json into Microsoft Excel file.
        self.material_json is like:
            {
                year_1:{
                    material_1:{
                        January:DataFrame,
                        Febuary:DataFrame,
                        ...
                    },
                    material_2:{
                        January:DataFrame,
                        Febuary:DataFrame,
                        ...
                    }
                },
                year_2:{
                    material_1:{
                        January:DataFrame,
                        Febuary:DataFrame,
                        ...
                    }
                }
            }
        '''
        from pandas import ExcelWriter
        from openpyxl import Workbook,load_workbook
        if os.path.isfile(savepath)==True:
            wb=load_workbook(savepath)
        elif os.path.isdir(savepath)==True:
            file_name=r'-'.join(['inventory',get_time_str(),r'.xlsx'])
            savepath=os.path.join(file_name,savepath)
            wb=Workbook()
            wb.save(savepath)
        else:
            wb=Workbook()
            wb.save(savepath)
        wter=ExcelWriter(savepath,engine='openpyxl')
        wter.book=wb
        pass
    @__time_monitor
    def start(self,export_path,moving_weighted_average=False):
        '''
<<<<<<< HEAD
        1.Load raw data;
        2.Calculate allocated amount of each product item;
        3.Get sum of cost for product items allocated;
        4.Write output data into Microsoft Excel, if necessary;
        5.Return sum of cost for product items.
>>>>>>> ab372ae28b6c4356792ebea1bb920191349a9cfa
        '''
        self.load_raw_data()
        self.cluster_material()
        self.cal_amount_allocated(moving_weighted_average=moving_weighted_average)
        self.export_material_data(export_path)
        resu=self.recal_product_cost()
        print('sum of allocated cost amount for each product','\n',resu)
        return resu
    pass
if __name__=='__main__':
    pass

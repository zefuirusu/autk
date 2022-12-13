#!/usr/bin/env python 
# coding=utf-8
'''
说明:
    1. AuSample类初始化时的部分属性没有得到充分利用,需要后期升级.
    2. AuSample.multiSample()和AuSample.getSample(inAcct)都是对金额抽样,而非对数量抽样,将来要更新一个对数量抽样的方法.
'''
import re
import threading

# from autk.financialtk.zhtk.zhchart import ChartAccount,Acct
# from autk.financialtk.zhtk.zhgl import Gele
# from autk.core.funcs import wtlog

from autk import MCA as ChartAccount
from autk.parser.entry import Acct,EntryRecord,JournalEntry,OppositeFinder
from autk import MGL as Gele
def wtlog():
    pass

class AuSample:
    '''
    gl_object is an instance of class Gele, and chart_object is an instance of class ChartAccount.
    String savedir is the directory of Sample Excel file to export.
    String acctli_dir is the input file indicating the list of account_id to sample.
    String logdir is the location of log file.
    Float acquired_rate is the target percent for sample size.
    List drcrdesc is the columns for [Debit_Amount,Credit_Amount] in Gele (General Ledger).
    无论是序时账还是余额表,发生额的名字都叫['借方','贷方'].
    '''
    def __init__(self,gl_object,chart_object,savedir,acctli_dir='./accountList.txt',logdir='./sampleLog.txt',acquired_rate=0.81,drcrdesc=[]):
        '''
        Get sample according to GL and TB.
        self.gl is an instance of class Gele and self.chart is an instance of class ChartAccount.
        parameters:(gl_object,chart_object,savedir,acctli_dir,logdir,acquired_rate,drcrdesc)
        self.savedir is actually savepath.
        To start, call method self.multiSample(single-thread) or self.start_multi_sample(multi-thread).
        余额表是['借方发生','贷方发生']
        序时账是['借方发生金额','贷方发生金额']
        '''
        self.gl=gl_object
        self.chart=chart_object
        self.savedir=savedir
        self.logdir=logdir
        def rm_space(instr):
            return self.rm_space(instr)
        if acctli_dir != None:
            with open(acctli_dir,mode='r',encoding='utf-8') as f:
                li=f.readlines()
                li=list(map(rm_space,li))
                li.sort(reverse=False) # reverse=False 升序.
                self.acctli=li
                pass
            print(self.acctli)
        self.acquired_rate=acquired_rate
        if  drcrdesc==[]:
            self.drcrdesc=self.gl.drcrdesc # r'借方发生金额',r'贷方发生金额'
        else:
            self.drcrdesc=drcrdesc
        print(self.acctli)
        self.thread_li=[]
        return
    def rm_space(self,instr):
        return re.sub(re.compile(r'\n.*$'),'',instr)
    def logw(self,logline):
        wtlog(logline,self.logdir)
        return
    def matsample(self,inAcct,mater_level=20000):
        '''
        get sample according to materiality.
        parameters:inAcct,mater_level
        '''
        logline='=='+str(inAcct.accid)+':'+str(inAcct.accna)+'=='
        print(logline)
        self.logw(logline)
        from pandas import DataFrame,Series
        theAcct=self.gl.accurate_filter(inAcct.accid,self.gl.accid_col) # 系统导出的序时账/余额表里是"科目编号",有时候从被审计单位系统导出的GL为“科目编码".
        # theAcct=self.gl.accurate_filter(inAcct.accid,r'科目编码') # 系统导出的序时账/余额表里是"科目编号",有时候从被审计单位系统导出的GL为“科目编码".
        theAcct=theAcct.fillna(0)
        acct_data=self.chart.getAcct(inAcct) # 从余额表获取发生额,用以计算抽样比例.
        # acct_sum=DataFrame({self.drcrdesc[0]:acct_data[2],self.drcrdesc[1]:acct_data[3]},index=[inAcct.accid],columns=self.drcrdesc)
        acct_sum=Series([acct_data[2],acct_data[3]],name=inAcct.accid,index=self.drcrdesc) # acct_sum是acct_data的一部分.
        # acct_sum=theAcct[self.drcrdesc].sum(axis=0) # 被指定的科目借方贷方求和.这里要修改,要从余额表读取此数.
        def sidesample(dcr):
            if acct_sum[dcr] != 0: # acct_sum[dcr] !=0的情况.
                if acct_sum[dcr]<0:
                    dcr_sample=theAcct[theAcct[dcr]<=-mater_level]
                    logline=dcr+' of %s is negative,so need attention.'%(inAcct.accid)
                    print(logline)
                else: # acct_sum[dcr]>0的情况.
                    dcr_sample=theAcct[theAcct[dcr]>=mater_level]
                    sam_rate=dcr_sample[dcr].sum(axis=0)/acct_sum[dcr]
                    logline=dcr+'sam_rate:%s'%sam_rate
                    print(logline)
                    self.logw(logline)
                if dcr_sample.shape[0]==0:
                    logline='no sample got for%s'%str(inAcct.accna+inAcct.accid)+dcr
                    self.logw(logline)
                    print(logline)
                else:
                    pass
            else: #acct_sum[dcr]==0:
                logline='%s :no need to sample.'%dcr
                print(logline)
                self.logw(logline)
                dcr_sample=DataFrame([],columns=theAcct.columns)
            return dcr_sample
        def getside():
            for i in self.drcrdesc:
                side=sidesample(i)
                yield side
        from pandas import concat
        final_sample=concat(getside(),axis=0,join='outer',ignore_index=True)
        final_sample=final_sample.reset_index(drop=True)
        self.logw('-----')
        return final_sample
    def getSample(self,inAcct):
        '''
        parameters:inAcct
        获取某科目的借方/贷方样本.
        '''
        from pandas import DataFrame,Series
        theAcct=self.gl.accurate_filter(inAcct.accid,self.gl.accid_col) # 系统导出的序时账/余额表里是"科目编号",有时候从被审计单位系统导出的GL为“科目编码".
        from numpy import float64
        theAcct[self.drcrdesc]=theAcct[self.drcrdesc].astype('float64',copy=True)
        # theAcct=self.gl.accurate_filter(inAcct.accid,r'科目编码') # 系统导出的序时账/余额表里是"科目编号",有时候从被审计单位系统导出的GL为“科目编码".
        # theAcct['借方发生金额'].dtype=float
        # theAcct['贷方发生金额'].dtype=float
        print('this acccout:')
        print(theAcct)
        acct_data=self.chart.getAcct(Acct(str(inAcct.accid))) # 从余额表获取发生额,用以计算抽样比例.
        print('acct_data:\n',acct_data)
        # acct_sum=DataFrame({self.drcrdesc[0]:acct_data[2],self.drcrdesc[1]:acct_data[3]},index=[inAcct.accid],columns=self.drcrdesc)
        acct_sum=Series([acct_data[2],acct_data[3]],name=inAcct.accid,index=self.drcrdesc) # acct_sum是acct_data的一部分.
        # acct_sum=theAcct[self.drcrdesc].sum(axis=0) # 被指定的科目借方贷方求和.这里要修改,要从余额表读取此数.
        sub_count=theAcct[self.drcrdesc].count(axis=0)
        target_sum=acct_sum*self.acquired_rate # 依据目标金额比例,确定目标累计合计金额.
        averAmount=acct_sum/sub_count # 平均每个样本的金额
        # start_nums=target_sum/averAmount # 计算初始样本容量,如acct_sum为0,start_nums会报错.
        logline=''.join([r'target_sum(Dr/Cr):','\n',str(target_sum[0]),'\t',str(target_sum[1])])
        self.logw(logline)
        print(logline)
        # print('target_sum:\n',target_sum)
        def loop_sample(dcr):
            '''
            对借方/贷方某方向抽样.
            dcr is one of ['借方','贷方']
            '''
            logline=''.join(['--',str(inAcct.accid),dcr,'--'])
            print(logline)
            self.logw(logline)
            if acct_sum[dcr] ==0:
                logline='%s :no need to sample.'%dcr
                print(logline)
                self.logw(logline)
                sloop=DataFrame([],columns=theAcct.columns)
                pass
            else: # acct_sum[dcr] !=0的情况.
                if acct_sum[dcr]<0:
                    logline=dcr+' of %s is negative,so need attention.'%(inAcct.accid)
                    print(logline)
                    self.logw(logline)
                    # Note: DataFrame.sample(n=None, frac=None, replace=False, weights=None, random_state=None, axis=None), frac为抽取的数量比例;replace=False,为无放回抽样;axis=0抽取列;random_state=None,取不重复数据.
                    sloop=theAcct.sample(frac=self.acquired_rate,replace=False,axis=0,random_state=None)
                    pass
                else: # acct_sum[dcr]>0的情况.
                    # n_start=int(target_sum[dcr]/averAmount[dcr]/2)
                    n_start=1
                    sloop=theAcct.nlargest(n=n_start,columns=[dcr],keep='last')
                    sam_rate=sloop[dcr].sum(axis=0)/acct_sum[dcr]
                    while sam_rate<self.acquired_rate:
                        n_start+=1
                        sloop=theAcct.nlargest(n=n_start,columns=[dcr],keep='last')
                        sam_rate=sloop[dcr].sum(axis=0)/acct_sum[dcr]
                        continue
                    logline=dcr+'sample_size:%d;'%n_start
                    self.logw(logline)
                    print(logline)
                    logline=''.join([dcr,'sam_rate:',str(sam_rate)])
                    self.logw(logline)
                    print(logline)
            return sloop
        def drcr_sample():
            '''
            分别抽借方/贷方.
            '''
            for i in self.drcrdesc:
                yield loop_sample(i)
        from pandas import concat
        final_sample=concat(drcr_sample(),axis=0,join='outer',ignore_index=True)
        final_sample=final_sample.reset_index(drop=True)
        return final_sample
    def multiSample(self):
        '''
        Single Thread and multi-account sample.
        '''
        print('%d accounts to sample in total.'%len(self.acctli))
        from openpyxl import load_workbook,Workbook
        from pandas import ExcelWriter
        wb=Workbook()
        wb.save(self.savedir)
        wb.close()
        wb=load_workbook(self.savedir)
        wter=ExcelWriter(self.savedir,engine='openpyxl')
        wter.book=wb
        self.logw('==multiSample==')
        # thread_list=[]
        for i in self.acctli:
            acct=Acct(i,self.chart.getna(i))
            # th=MultiThread(acct,self)
            # thread_list.append(th)
            print('==start:%s=='%str(acct.accid))
            self.logw('==start:%s=='%acct.accid)
            m_sample=self.getSample(acct) # one of the multi-samples.
            # try:
            #     m_sample=self.getSample(acct) # one of the multi-samples.
            # except:
            #     from pandas import DataFrame
            #     m_sample=self.getSample(acct) # one of the multi-samples.
            #     # m_sample=
            #     print('sample for this account failed:')
            #     print(acct.accid,'\t',acct.accna)
            #     logline='\t'.join([acct.accid,acct.accna])
            #     self.logw(logline)
            #     self.logw('sample for this account failed:')
            #     lgline=acct.accid+'\t'+acct.accna
            #     self.logw(lgline)
            #     pass
            if m_sample.shape[0]==0:
                self.logw('no sample for this account:')
                no_sample_line=''.join([str(acct.accid),'\t',str(acct.accna)])
                self.logw(no_sample_line)
                pass
            else:
                m_sample.to_excel(wter,sheet_name=str(acct.accid+acct.accna))
                # yield list(m_sample.loc[:,'glid'].drop_duplicates())
                # print(m_sample) # 查看样本
                wter.save()
            # yield m_sample
            print('==end:%s=='%str(acct.accid))
            self.logw('==end:%s=='%acct.accid)
        wter.close()
        return
    def start_multi_sample(self,trans_accid=True):
        '''
        Multi-Thread sample.
        '''
        self.logw('multi-thread multiSample start....')
        print('multi-thread multiSample start....')
        print('%d accounts to sample in total.'%len(self.acctli))
        from openpyxl import load_workbook,Workbook
        from pandas import ExcelWriter
        wb=Workbook()
        wb.save(self.savedir)
        wb.close()
        wb=load_workbook(self.savedir)
        wter=ExcelWriter(self.savedir,engine='openpyxl')
        wter.book=wb
        self.logw('==multiSample==')
        thread_list=[]
        for i in self.acctli:
            acct=Acct(i,self.chart.getna(i,trans_accid))
            print('account:',acct.__dict__)
            th=MultiThread(acct,self,wter)
            thread_list.append(th)
            continue
        for i in thread_list:
            i.start()
        for i in thread_list:
            i.join()
        wter.save()
class MultiThread(threading.Thread):
    def __init__(self,acct,sample_obj,wter):
        self.acct=acct
        self.sample_obj=sample_obj
        self.wter=wter
        threading.Thread.__init__(self,name=self.acct)
        pass
    def run(self):
        print('==start:%s=='%str(self.acct.accid))
        self.sample_obj.logw('==start:%s=='%self.acct.accid)
        m_sample=self.sample_obj.getSample(self.acct) # one of the multi-samples.
        if m_sample.shape[0]==0:
            self.sample_obj.logw('no sample for this account:')
            no_sample_line=''.join([str(self.acct.accid),'\t',str(self.acct.accna)])
            self.sample_obj.logw(no_sample_line)
            pass
        else:
            m_sample.to_excel(self.wter,sheet_name=''.join([str(self.acct.accid),str(self.acct.accna)]))
            # self.wter.save()
        print('==end:',str(self.acct.accid),'==')
        self.sample_obj.logw('==end:%s=='%str(self.acct.accid))
        pass
class genSample:
    def __init__(self,gldir,shtna,title=0,method='pm'):
        self.gldir=gldir
        self.sheetname=shtna
        self.title=title
        self.method=method
    def iterate_data(self):
        '''
        iterate rows and return EntryRecord data.
        '''
        from pandas import read_excel
        # from autk.financialtk.journal import EntryRecord
    pass
if __name__=='__main__':
    pass

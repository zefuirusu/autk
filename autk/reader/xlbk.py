#!/usr/bin/env python
# -*- coding: utf-8 -*-

import re
import os
from xlrd import open_workbook
from openpyxl import load_workbook
from pandas import read_excel,DataFrame
class XlBook:
    '''
    Basic Structure of XlBook on default:
        file_path,
        file_name,
        suffix:xls/xlsx/xlsm,
        shtli,
        data,
    '''
    def __init__(self,file_path):
        self.file_path=file_path
        self.file_name=''
        self.pure_file_name=''
        self.suffix=''
        #  self.shtli=[]
        self.__parse_file_type()
        pass
    def __parse_file_type(self):
        self.file_name=self.file_path.split(os.sep)[-1]
        self.pure_file_name=re.sub(
            re.compile(r'\.xls[xm]?'),
            '',
            self.file_name
        )
        self.suffix=re.sub(
            re.compile(self.pure_file_name+r'.'),
            '',
            self.file_name
        )
        pass
    @property
    def shtli(self):
        if self.suffix=='xls':
            return open_workbook(self.file_path).sheets()
        elif self.suffix=='xlsx':
            return load_workbook(self.file_path).sheetnames
        elif self.suffix=='xlsm':
            return load_workbook(self.file_path,keep_vba=True).sheetnames
        pass
    def find_sheet(self,regex_str):
        from autk.parser.funcs import regex_filter
        possible_names=regex_filter(regex_str,self.shtli,match_mode=False)
        return possible_names
    @property
    def shape(self):
        '''
        sheet_name | shape
        -----------|-------
        sheet_1    | (n_1,m_1)
        sheet_2    | (n_2,m_2)
        .......
        '''
        from numpy import array
        if self.suffix=='xls':
            shape=array(
                [
                    [open_workbook(self.file_path).sheet_by_name(sht).nrows,
                     open_workbook(self.file_path).sheet_by_name(sht).ncols]
                    for sht in self.shtli
                ]
            )
            pass
        elif self.suffix=='xlsx' or 'xlsm':
            shape=array(
                [
                    [load_workbook(self.file_path)[sht].max_row,
                     load_workbook(self.file_path)[sht].max_column]
                    for sht in self.shtli
                ]
            )
            pass
        return shape
    @property
    def shape_df(self):
        from pandas import DataFrame
        return DataFrame(
            data=self.shape,
            columns=['rows','cols'],
            index=self.shtli
        )
    def get_matrix(
            self,
            sheet_name,
            start_cell_index,
            n_rows_range,
            n_cols_range,
            type_df=False,
            has_title=False
    ):
        '''
        start_cell_index is a tuple like 'R1C1' ref-style in Excel: (row,column);
        For xlrd.open_workbook().sheet_by_name(), index starts from 0;
        While for openpyxl.load_workbook().get_sheet_by_name(), index starts from 1;
        That's all right, just start from 1 when passing argument 'start_cell_index' as tuple like (n,m).
        For numbers, different file type results in different data type:
            xls:str(float)
            xlsx/xlsm:str(int)
        '''
        from numpy import array
        if self.suffix=='xls':
            sht=open_workbook(self.file_path).sheet_by_name(sheet_name)
            matrix=array(
                [
                    sht.row_values(
                        row,
                        start_cell_index[1]-1,
                        start_cell_index[1]-1+n_cols_range
                    ) for row in range(
                        start_cell_index[0]-1,
                        start_cell_index[0]-1+n_rows_range
                    )
                ]
            )
        elif self.suffix=='xlsx':
            #  sht=load_workbook(self.file_path).get_sheet_by_name(sheet_name) # same as:
            sht=load_workbook(self.file_path)[sheet_name]
            matrix=array(
                list(
                    sht.iter_rows(
                        min_row=start_cell_index[0],
                        max_row=start_cell_index[0]+n_rows_range-1,
                        min_col=start_cell_index[1],
                        max_col=start_cell_index[1]+n_cols_range-1,
                        values_only=True
                    )
                )
            )
        elif self.suffix=='xlsm':
            #  sht=load_workbook(self.file_path).get_sheet_by_name(sheet_name) # same as:
            sht=load_workbook(self.file_path,keep_vba=True)[sheet_name]
            matrix=array(
                list(
                    sht.iter_rows(
                        min_row=start_cell_index[0],
                        max_row=start_cell_index[0]+n_rows_range-1,
                        min_col=start_cell_index[1],
                        max_col=start_cell_index[1]+n_cols_range-1,
                        values_only=True
                    )
                )
            )
        else:
            from numpy import zeros
            matrix=zeros((n_rows_range,n_cols_range))
        if type_df==True:
            if has_title==False:
                matrix=DataFrame(
                    data=matrix
                )
            else:
                matrix=DataFrame(
                    data=matrix[1:],
                    columns=matrix[0]
                )
        else:
            pass
        return matrix
    def select_matrix(
        self,
        sheet_name,
        from_cell_index,
        to_cell_index,
        type_df=False,
        has_title=False
    ):
        '''
        from_cell_index and to_cell_index are tuples like 'R1C1' ref-style in Excel: (row,column);
        '''
        return self.get_matrix(
            sheet_name,
            from_cell_index,
            to_cell_index[0]-from_cell_index[0]+1,
            to_cell_index[1]-from_cell_index[1]+1,
            type_df=type_df,
            has_title=has_title
        )
    def get_row(self,sheet_name,row):
        max_col=self.shape_df.at[sheet_name,'cols']
        return list(
            self.select_matrix(
                sheet_name,
                (row,1),
                (row,max_col),
                type_df=False,
                has_title=False
            )[0]
        )
    def get_col(self,sheet_name,col):
        max_row=self.shape_df.at[sheet_name,'rows']
        return list(
            self.select_matrix(
                sheet_name,
                (1,col),
                (max_row,col),
                type_df=False,
                has_title=False
            )[0]
        )
    def test_map(self,xlmap,common_title=0):
        '''
        To test each sheet to check if they are fit to the input `xlmap`;
        '''
        resu_df=DataFrame([],index=self.shape_df.index,columns=xlmap.columns)
        map_cols=xlmap.columns
        map_dict=xlmap.show
        for sht in resu_df.index:
            max_cols=self.shape_df.at[sht,'cols']
            sht_cols=self.get_row(sht,1)
            for col in resu_df.columns:
                col_index=map_dict[col]
                if col_index is not None:
                    resu_df.at[sht,col]=sht_cols[col_index]
                continue
            continue
        return resu_df
    def to_mtb(self,common_title=0,auto_load=False):
        '''
        Transform self into ImmortalTable.
        '''
        from autk.reader.table import ImmortalTable
        xlmeta={}
        xlmeta.update(
            {self.file_path:[[sht,common_title] for sht in
                             self.shtli]}
        )
        return ImmortalTable(
            xlmeta=xlmeta,
            common_title=common_title,
            xlmap=None,
            use_map=False,
            auto_load=auto_load,
            keep_meta_info=True,
            #  key_index=[],
            #  key_name='key_id'
        )
        pass
    def to_mgl(
        self,
        common_title=0,
        xlmap=None,
        auto_load=False
    ):
        from autk.reader.mortalgl import MGL
        xlmeta={}
        xlmeta.update(
            {self.file_path:[[sht,common_title] for sht in
                             self.shtli]}
        )
        return MGL(
            xlmeta=xlmeta,
            common_title=common_title,
            xlmap=xlmap,
            auto_load=auto_load,
            nick_name='mgl_frbk'
        )
    def to_chart(
        self,
        common_title=0,
        xlmap=None,
        auto_load=False
    ):
        from autk.reader.chart import MCA
        xlmeta={}
        xlmeta.update(
            {self.file_path:[[sht,common_title] for sht in 
                             self.shtli]}
        )
        return MCA(
            xlmeta=xlmeta,
            common_title=common_title,
            xlmap=xlmap,
            auto_load=auto_load,
            key_cols=[],
            nick_name='mca_frbk'
        )
        pass
    def to_inventory(
        self,
    ):
        pass
    pass
if __name__=='__main__':
    pass

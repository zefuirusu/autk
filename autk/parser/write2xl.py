#!/usr/bin/env python
# encoding = 'utf-8'
'''
Seems useless.
'''
from os import sep
from os.path import isfile,isdir
from os.path import join as osjoin
from pandas import ExcelWriter
from openpyxl import load_workbook,Workbook
from autk import get_time_str
class ThWter:
    def __init__(self,save_path,nick_name='data'):
        self.wter=None
        self.file_name=None
        self.save_path=None
        self.nick_name=nick_name
        self.__parse_path(save_path)
        pass
    def __parse_path(self,save_path):
        if isfile(save_path):
            self.save_path=save_path
            self.wter=ExcelWriter(save_path)
            self.file_name=save_path.split(sep)
            wb=load_workbook(save_path)
            self.wter.book=wb
        elif isdir(save_path):
            file_name=''.join([self.nick_name,get_time_str(woc=True),r'.xlsx'])
            self.file_name=file_name
            self.save_path=osjoin(save_path,self.file_name)
            wb=Workbook()
            wb.save(save_path)
            self.wter=ExcelWriter(save_path)
            self.wter.book=wb
        elif isinstance(self.save_path,str):
            wb=Workbook()
            wb.save(self.save_path)
            self.wter=ExcelWriter(save_path)
            self.wter.book=wb
            pass
        else:
            pass
        pass
    def save_df(self,in_df,sheet_name=None):
        if sheet_name is None:
            sheet_name=get_time_str(woc=True)
        self.__parse_path(self.save_path)
        in_df.to_excel(self.wter,sheet_name=sheet_name)
        self.wter.save()
        pass
    pass
if __name__=='__main__':
    pass

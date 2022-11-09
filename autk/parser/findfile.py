#!/usr/bin/env python
# coding=utf-8
import re
import os
import datetime
from threading import Thread
from autk.parser.funcs import get_time_str
class Walker:
    def __init__(
            self,
            item,
            search_dir=os.path.abspath(os.curdir),
            match=False
        ):
        self.search_dir=search_dir
        self.match=match
        self.item=item
        self.regitem=re.compile(self.item)
        self.resu_files=None
        self.resu_dirs=None
        pass
    def one_match(self,compare_target):
        if self.match==False:
            return re.search(self.regitem,compare_target)
        else:
            return re.match(self.regitem,compare_target)
        pass
    def get_files(self):
        for i,j,k in os.walk(self.search_dir):
            for file_name in k:
                if self.one_match(file_name) is not None:
                    yield os.path.abspath(os.path.join(i,file_name))
        pass
    def get_dirs(self):
        for i,j,k in os.walk(self.search_dir):
            for file_path in j:
                if self.one_match(file_path) is not None:
                    yield os.path.abspath(file_path)
        pass
    def start_search(self):
        def th1():
            self.resu_files=self.get_files()
        def th2():
            self.resu_dirs=self.get_dirs()
        t1=Thread(target=th1)
        t2=Thread(target=th2)
        t1.start()
        t1.join()
        t2.start()
        t2.join()
        pass
    pass
def add_nick_name(nick_name,suf_type,file_dir):
    import shutil
    def __change_file_name(file_path):
        new_file_path_list=file_path.split(os.sep)
        new_file_path_list[-1]=re.sub(r'\.'+suf_type+'$',r'-'+nick_name+r'.'+suf_type,file_path.split(os.sep)[-1])
        new_file_path=os.sep.join(
            new_file_path_list
        )
        print('\nold_name:\n',
              file_path.split(os.sep)[-1],
              '\n',
              'new_name:\n',
              new_file_path.split(os.sep)[-1]
        )
        shutil.move(file_path,new_file_path)
        pass
    w1=Walker(r'\.'+suf_type+r'$',file_dir,match=False)
    # map(__change_file_name,w1.get_files())
    for f in w1.get_files():
        __change_file_name(f)
        continue
    pass
def find_regex(item,search_dir=os.path.abspath(os.curdir),match=False):
    '''
    To find and return `files/paths` according to regex `item` given, in `search_dir`.
    Search mode/Match mode are both supported.
    '''
    print('start time:',datetime.datetime.now())
    w1=Walker(item=item,search_dir=search_dir,match=match)
    w1.start_search()
    resu=[list(w1.resu_files),list(w1.resu_dirs)]
    print('end time:',datetime.datetime.now())
    return resu
def scanxl(xl_dir=os.path.abspath(os.curdir)):
    from openpyxl import load_workbook
    from pandas import read_excel
    f=find_regex(r'\.xlsx',xl_dir)[0]
    for i in f:
        print('='*5)
        print('bookName:',i.split(os.sep)[-1])
        shtli=load_workbook(i).sheetnames
        print('sheets:',shtli)
        for j in shtli:
            print('\t','sheetName:',j)
            print('\t'*2,'sheetCols:',read_excel(i,sheet_name=j,engine='openpyxl').columns)
        print('-'*5)
        continue
    pass
def gendirs(target_dir):
    '''
    This function seems useless now.
    make directories according to the name of files in the current directory.
    去掉所有的中文字符,以英文名字创建文件夹,用以整理资料.
    '''
    for file_name in os.listdir(path=target_dir):
        pure_file_name=re.sub(r'\.pdf$',r'',file_name)
        no_zh_dir_name=re.sub(r'-*[\u4e00-\u9fa5]+.+',r'',pure_file_name)
        print('replace 1: %s'%pure_file_name)
        print('replace 2: %s'%no_zh_dir_name)
        try:
            os.makedirs(no_zh_dir_name)
        except FileExistsError:
            print('%s exists!'%no_zh_dir_name)
            pass
def wcp(
        item,
        frdir,
        todir=os.path.abspath(os.curdir),
        nickName=get_time_str()
):
    '''
    Only files, not directories can be copied.
    frdir is the 'from and search' directory, todir is the 'to' directory.
    Easy copy function using with the find function and then yelling out a big wocao.
    Works well with the find function when you need to copy all the files that you find to a target new directory.
    '''
    import shutil
    cp_files=find_regex(item,frdir)[0]
    file_name='-'.join([nickName,str(frdir.split(os.sep)[-1])])
    destination_dir=os.path.abspath(os.path.join(todir,file_name))
    try:
        shutil.copy(frdir,destination_dir)
    except:
        shutil.copytree(frdir,destination_dir)
    else:
        pass
    finally:
        print('%s is copied.'%frdir)
    return
#
if __name__=='__main__':
    pass

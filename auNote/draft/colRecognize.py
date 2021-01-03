#!/usr/bin/env python
# coding=utf-8
'''
Column Recognizer.
Read columns of DataFrame and fit them with regular expression,
'''
class Reco:
    def __init__(self):
        super().__init__()
        pass
    def loadcols(self,xldir):
        from pandas import read_excel
        from openpyxl import load_workbook
        shtli=load_workbook(xldir).sheetnames
        df_part=read_excel(xldir,sheet_name=shtli[0],header=3,engine='openpyxl')
        df_part=df_part.head(15)
        print(shtli)
        print(df_part)
        # print(df_part.columns)
        print(df_part.values)
        # nan=df_part.iloc[0,0]
        # print(nan)
        # print(type(nan))
        return
#
if __name__=='__main__':
    testdir='./bjXinLeiNeng-GL-1-10月-output2020-1228.xlsx'
    r1=Reco()
    r1.loadcols(testdir)

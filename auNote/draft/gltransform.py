#!/usr/bin/env python
# coding=utf-8
from openpyxl import load_workbook
from pandas import concat
from autk.financialtk.zhgl import Gele
gldir='./bjgl-1-10月-output2020-1228.xlsx'
savedir='/mnt/e/skandha/z-Sync/a-Project/xin新雷能/glandtb/北京/yield_gl.xlsx'
shtli=load_workbook(gldir).sheetnames
# print(shtli)
'''
['表页-1', '表页-2', '表页-3', '表页-4', '表页-5', '表页-6', '表页-7', '表页-8', '表页-9', '表页-10', '表页-11', '表页-12', '表页-13', '表页-14', '表页-15', '表页-16', '表页-17', '表页-18', '表页-19', '表页-20']

'''
def get_gls():
    for i in shtli:
        gl=Gele(gldir,i)
        yield gl.getgldata()
    pass
def main():
    df=concat(get_gls(),axis=0,join='outer')
    df.to_excel(savedir)
if __name__=='__main__':
    main()
    pass

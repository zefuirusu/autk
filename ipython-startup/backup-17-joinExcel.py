#!/usr/bin/env python
# -*- coding: utf-8 -*-
import os
import re
import openpyxl as ox
<<<<<<< HEAD
from find import find
from pandas.core.frame import DataFrame
#
def compareli(li1=[],li2=[]):
    '''
    li1 and li2 are both list-type.验证科目列表是否一致
    resu=[shared,li1priv,li2priv].
    '''
    shared=[]
    li1priv=[]
    li2priv=[]
    for i in li1:
        if i in li2:
            shared.append(i)
        else:
            li1priv.append(i)
    for i in li2:
        if i in li1:
            if i in shared:
                pass
            else:
                shared.append(i)
        else:
            li2priv.append(i)
    resu=[shared,li1priv,li2priv]
    return resu
=======
from 11-skfind import find
from pandas.core.frame import DataFrame
#
# def compareli(li1=[],li2=[]):
#     '''
#     li1 and li2 are both list-type.验证科目列表是否一致
#     resu=[shared,li1priv,li2priv].
#     '''
#     shared=[]
#     li1priv=[]
#     li2priv=[]
#     for i in li1:
#         if i in li2:
#             shared.append(i)
#         else:
#             li1priv.append(i)
#     for i in li2:
#         if i in li1:
#             if i in shared:
#                 pass
#             else:
#                 shared.append(i)
#         else:
#             li2priv.append(i)
#     resu=[shared,li1priv,li2priv]
#     return resu
>>>>>>> 9535ed2f76671bcf590473cbf1e221df7fcc5b03
#
def prejoin():
    f=find(r'\.xlsx',os.path.abspath(os.curdir))[0]
    for i in f:
        print('='*5)
        print('bookName:',i.split(os.sep)[-1])
        shtli=ox.load_workbook(i).sheetnames
        print('sheets:',shtli)
        for j in shtli:
            print('\t','sheetName:',j)
            print('\t'*2,'sheetCols:',pd.read_excel(i,sheet_name=j).columns)
        print('-'*5)
#
def joinsheet(indir,outdir=os.path.abspath(os.curdir),shtli=None,keep=True,title=0,fill=False):
    '''
    Join all sheets of one Excel book at 'indir' into a single sheet at 'outdir' and will tell you 'fromWhichSheet'.
    The input file has to be 'xlsx' format!
    'indir' and 'outdir' are two paths.
    'outdir' is default to the current directory.
    If 'keep' is True, the output file will add a column to show the original sheetName.
    Also return the output pandas.DataFrame object.
    '''
    if shtli==None:
        import openpyxl as ox
        shtli=ox.load_workbook(indir).sheetnames
    else:
        pass
    import pandas as pd
    d=pd.DataFrame([])
    for i in shtli:
        d_fake=pd.read_excel(indir,sheet_name=i,header=title)
        if keep==True:
            d_fake['fromWhichSheet']=[str(i)]*d_fake.shape[0]
        else:
            pass
        if fill==True:
            d_fake=d_fake.fillna(method='pad')
        else:
            pass
        d_fake=d_fake.drop_duplicates().reset_index(drop=True)
        d=pd.concat([d,d_fake],join='outer',axis=0)
    #
    outfilename=re.sub(r'\.xlsx$',r'-joinsheet.xlsx',indir.split(os.sep)[-1])
    outPath=os.path.join(outdir,outfilename)
    # d=d.fillna(method='ffill')
    # d=d.drop_duplicates().reset_index(drop=True)
    d.to_excel(outPath)
    print('input file sheet list:','\n',shtli)
    print('output file: %s'%outPath)
    print('output file shape:\n',d.shape)
    print('output file columns:','\n',d.columns)
    return d
#
def joinsheetplus(bookli=os.listdir(os.path.abspath(os.curdir)),outdir=os.path.abspath(os.curdir),shtna=None,title=0,keep=True,fill=False):
    '''
    Join sheets from different Excel books into a single Sheet.
    And will tell you 'fromWhichBook'.
    '''

    outfilename=outdir+r'-joinsheetplus.xlsx'
    import pandas as pd
    d=pd.DataFrame([])
    for i in bookli:
        dir_fake=os.path.abspath(i)
        if shtna==None:
            d_fake=pd.read_excel(dir_fake,header=title)
        else:
            d_fake=pd.read_excel(dir_fake,sheet_name=shtna,header=title)
        if keep==True:
            d_fake['fromWhichBook']=[str(i)]*d_fake.shape[0]
        else:
            pass
        if fill==True:
            d_fake=d_fake.fillna(method='pad')
        else:
            pass
        d_fake=d_fake.drop_duplicates().reset_index(drop=True)
        d=pd.concat([d,d_fake],join='outer',axis=0)
    d.to_excel(outfilename)
    print('input file list:','\n',bookli)
    print('output file location:','\n',outfilename)
    print('output file shape:','\n',d.shape)
    print('outpit file columns:','\n',d.columns)
    return d
#
def joinbook(bookli=os.listdir(os.path.abspath(os.curdir)),outdir=os.path.abspath(os.curdir),shtna=None,title=0):
    '''
    Join sheets from different books(in the same position) into one single book, well, in different sheets.
    '''
    outPath=outdir+r'-joinbook.xlsx'
<<<<<<< HEAD
=======
    import pandas as pd
>>>>>>> 9535ed2f76671bcf590473cbf1e221df7fcc5b03
    from pandas import ExcelWriter
    wter=ExcelWriter(outPath)
    bkli=[]
    for i in bookli:
        if re.search(r'.*\.xlsx$',i) != None:
            bkli.append(i)
        else:
            pass
    for i in bkli:
<<<<<<< HEAD
        b=pd.read_excel(i,sheet_name=shtna,title=0)
        b.to_excel(wter,sheet_name=str(i))
=======
        if shtna != None:
            b=pd.read_excel(i,sheet_name=shtna,header=title)
            b.to_excel(wter,sheet_name=str(i))
        else:
            b=pd.read_excel(i,header=title)
            b.to_excel(wter,sheet_name=str(i))
>>>>>>> 9535ed2f76671bcf590473cbf1e221df7fcc5b03
    wter.save()
#

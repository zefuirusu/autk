#!/usr/bin/env python
# coding=utf-8
import re
import os
from openpyxl import load_workbook
from pandas import read_excel,concat
from autk.fileAssistant.findfile import find
class JoinExcel:
    def __init__(self,inpath,title=3,keep=True,fill=False,savedir=os.path.abspath(os.curdir)):
        self.inpath=inpath
        self.flij=find(r'\.xlsx',self.inpath)[0] # file list to join
        self.title=title
        self.keep=keep
        self.fill=fill
        self.savedir=savedir
        pass
    def prejoin(self):
        self.metadata={}
        for i in self.flij:
            print('='*5)
            print('bookName:',i.split(os.sep)[-1])
            shtli=load_workbook(i).sheetnames
            print('sheets:',shtli)
            for j in shtli:
                print('\t','sheetName:',j)
                print('\t'*2,'sheetCols:',read_excel(i,sheet_name=j,header=3,engine='openpyxl').columns)
                print('-'*5)
        pass
    def joingl(self):
        pass
    @staticmethod
    def joinsheet(inbook,shtli=None,title=3,save=False,keep=True,fill=False):
        if shtli==None:
            shtli=load_workbook(inbook).sheetnames
            pass
        else:
            pass
        pass
        def getfli():
            for i in shtli:
                d_fake=read_excel(inook,sheet_name=i,header=title)
                if keep==True:
                    d_fake['fromWhichSheet']=[str(i)]*d_fake.shape[0]
                    pass
                else:
                    pass
                if fill==True:
                    d_fake=d_fake.fillna(method='pad')
                    pass
                else:
                    pass
                d_fake=d_fake.drop_duplicates().reset_index(drop=True)
                yield d_fake
        d=concat(getfli(),join='outer',axis=0)
        if save==False:
            return d
        else:
            outfilename=re.sub(r'\.xlsx$',r'-joinsheet.xlsx',inbook)
            d.to_excel(outfilename)
            return
    def joinsheetplus(self,shtna=None,title=3,keep=True,fill=False,savedir=os.path.abspath(os.curdir)):
        '''
        Join sheets from different Excel books into a single Sheet.
        And will tell you 'fromWhichBook'.
        '''
        def getfli():
            for i in self.inpath:
                if shtna==None:
                    d_fake=read_excel(i,header=self.title)
                else:
                    d_fake=read_excel(i,sheet_name=shtna,header=self.title)
                if keep==self.keep:
                    d_fake['fromWhichBook']=[str(i)]*d_fake.shape[0]
                    pass
                else: # keep==False
                    pass
                if fill==self.fill:
                    d_fake=d_fake.fillna(method='pad')
                    pass
                else: # fill==False
                    pass
                d_fake=d_fake.drop_duplicates().reset_index(drop=True)
                yield d_fake
        d=concat(getfli(),join='outer',axis=0)
        if savedir != None:
            outfilename=savedir+r'-joinsheetplus.xlsx'
            d.to_excel(outfilename)
            return d
        else:
            return d
    def joinbook(self,shtna=None,title=3,keep=True,fill=False,savedir=os.path.abspath(os.curdir)):
        '''
        Join sheets from different books(in the same position) into one single book, well, in different sheets.
        '''
        def getfli():
            for i in self.inpath:
                if shtna==None:
                    d_fake=read_excel(i,header=self.title)
                else:
                    d_fake=read_excel(i,sheet_name=shtna,header=self.title)
                if keep==self.keep:
                    d_fake['fromWhichBook']=[str(i)]*d_fake.shape[0]
                    pass
                else: # keep==False
                    pass
                if fill==self.fill:
                    d_fake=d_fake.fillna(method='pad')
                    pass
                else: # fill==False
                    pass
                d_fake=d_fake.drop_duplicates().reset_index(drop=True)
                yield d_fake
        from pandas import ExcelWriter
        wter=ExcelWriter()
        for i in getfli():
            i.to_excel(wter,sheet_name=str(i),engine='openpyxl')
        wter.save()
        return

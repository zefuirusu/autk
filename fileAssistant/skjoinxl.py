#!/usr/bin/env python
# coding=utf-8
'''
Join Excel sheets.
'''
import os 
import threading
from openpyxl import load_workbook,Workbook
from pandas import read_excel,concat,ExcelWriter
def get_time_str(woc=False):
    import time
    time_list=list(map(str,time.localtime()))
    if len(time_list[1])==1:
        time_list[1]='0'+time_list[1]
        pass
    else:
        pass
    if woc == False:
        timestr='-'.join(['-'.join(time_list[0:3]),'_'.join(time_list[3:6])])
    else:
        timestr=''.join(time_list[0:6])
    return timestr
class TarSheet:
    def __init__(self,bookpath,sheetname,title):
        '''
        Pass arguments of sheetname and title.
        '''
        self.bookpath=bookpath
        self.sheetname=sheetname
        self.title=title
        self.data=None
        pass
    def read(self):
        d=read_excel(self.bookpath,sheet_name=self.sheetname,header=self.title,engine='openpyxl')
        self.data=d
        print(d.shape)
        # return d
class MultiRead(threading.Thread):
    def __init__(self,tar_sheet):
        self.data=None
        self.tar_sheet=tar_sheet
        self.bookpath=tar_sheet.bookpath
        self.sheetname=tar_sheet.sheetname
        self.title=tar_sheet.title
        self.thread_name=r':'.join([self.bookpath.split(os.sep)[-1],self.sheetname])
        threading.Thread.__init__(self,name=self.thread_name)
        pass
    def run(self):
        print(self.bookpath)
        print(threading.current_thread().name)
        self.tar_sheet.read()
        self.data=self.tar_sheet.data
        pass
class JoinExcel:
    def __init__(self,savedir=os.path.abspath(os.curdir),prefix='',suffix=get_time_str(woc=True)):
        self.savedir=savedir
        self.prefix=prefix
        self.suffix=suffix
        self.tar_sheet_list=None
        self.thread_list=[]
        self.data_list=[]
        pass
    def get_tarsheets(tar_dir,shtna='Sheet1',title=0):
        '''
        Same sheetname, same title_index.
        '''
        def yield_tars():
            for i in os.listdir(tar_dir):
                i_fake=[os.path.abspath(os.path.join(tar_dir,i)),shtna,title]
                ts=TarSheet(i_fake[0],i_fake[1],i_fake[2])
                yield ts
                continue
            pass
        self.tar_sheet_list=list(yield_tars())
        return self.tar_sheet_list
    def load_sheets(self,tar_sheet_list):
    '''
    Target Sheet list must be instance of class TarSheet.
    '''
        self.tar_sheet_list=tar_sheet_list
        pass
    def read_tarsht(self):
        for tar_sheet in self.tar_sheet_list:
            mr=MultiRead(tar_sheet)
            self.thread_list.append(mr)
            # d=read_excel(tar_sheet.bookpath,sheet_name=tar_sheet.sheetname,header=tar_sheet.title,engine='openpyxl')
            # mr.start()
            pass
        print(self.thread_list)
        for i in self.thread_list:
            i.start()
            i.join()
            self.data_list.append(i.data)
    def join_to_sheet(self,bookname='join',sheetname='join',write=False):
        '''
        parameters:
            bookname:
                Name of the export Excel Workbook.
            sheetname:
                Name of the Sheet in the exporting Excel Workbook.
        '''
        savename=''.join([self.prefix,bookname,self.suffix,r'.xlsx'])
        savepath=os.path.join(self.savedir,savename)
        self.read_tarsht()
        d=concat(self.data_list,axis=0,join='outer')
        if write == True:
            d.to_excel(savepath)
            pass
        else:
            return d
    def join_to_book(self,bookname='join',split=True,write=False):
        '''
        parameters:
            bookname:
                Name of the export Excel.
            split:
                If true,save to different sheets of the Excel Workbook.
                Else, save to single sheet of the Excel Workbook.
        '''
        savename=''.join([self.prefix,bookname,self.suffix,r'.xlsx'])
        savepath=os.path.join(self.savedir,savename)
        self.read_tarsht()
        if split == True:
            n=1
            for i in self.data_list:
                if write==True:
                    wb=Workbook()
                    wb.save(savepath)
                    wb.close()
                    wter=ExcelWriter(path=savepath,engine='openpyxl')
                    wter.book=wb
                    i.to_excel(wter,sheet_name=str(n))
                    pass
                else:
                    print('cannot show multi sheets.')
                    return
                n+=1
                continue
            pass
        else:
            d=concat(self.data_list,axis=0,join='outer')
            if write == True:
                wb=Workbook()
                wb.save(savepath)
                wb.close()
                wter=ExcelWriter(path=savepath,engine='openpyxl')
                wter.book=wb
                d.to_excel(wter,sheet_name='join_to_book')
                pass
            else:
                return d
        pass
if __name__=='__main__':
    print(get_time_str(woc=True))
    print(get_time_str(woc=False))
    pass

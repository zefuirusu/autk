#!/usr/bin/env python
# coding=utf-8
'''
Draw records of specific Account.
Draw Other Income/Expense from 7 Excel files of General Ledger.
basedir is where GL Excel files locates.
accid_li determines which accounts you are going to draw.
glid_index is default to ['账簿编码','期间年','期间月','凭证编码']
'''
import os
import re
import threading
from pandas import concat,ExcelWriter
from openpyxl import load_workbook,Workbook
from autk.financialtk.modules.mortalGL import MGL
basedir='../gl2020'
accid_li=[
    '6112' # 补贴收入
    # '6711',
    # '6301'
]
# accid_reg_li=[
#     r'^6711.*',
#     r'^6301.*'
# ]
def trans_str(instr):
    '''
    Transform and compile common string into regular expression items.
    '''
    return str(''.join([r'^',instr,r'.*']))
def get_files(basedir):
    '''
    Read Excel files.
    '''
    for i in os.listdir(basedir):
        yield os.path.abspath(os.path.join(basedir,i))
fileli=list(get_files(basedir))
def test():
    m1=MGL(fileli[0],'Sheet1',0)
    d1=get_record(accid_li[0],m1)
    print(d1)
    pass
def get_record(accid_li,mgl,savedir='../output/qtsz'):
    savename=''.join(['sample-qtsz-',mgl.fpath.split(os.sep)[-1]])
    savepath=os.path.abspath(os.path.join(savedir,savename))
    # wb=load_workbook(savepath)
    if os.path.exists(savepath) == True:
        print('exists!%s'%savepath)
        wb=load_workbook(savepath)
    else:
        wb=Workbook()
        wb.save(savepath)
        wb.close()
    wter=ExcelWriter(savepath,engine='openpyxl')
    wter.book=wb
    mgl.load_raw_data()
    mgl.set_glid()
    # mgl.set_glid(['账簿编码','期间年','期间月','凭证编码'])
    for accid in accid_li:
        data=mgl.getAcct(trans_str(accid),pure=True)
        data.to_excel(wter,sheet_name=accid,engine='openpyxl')
        wter.save()
    # return data
    pass
class GetGLThread(threading.Thread):
    def __init__(self,accid_li,mgl,savedir='../output/qtsz'):
        # self.accid=accid
        self.accid_li=accid_li
        self.mgl=mgl
        self.savedir=savedir
        threading.Thread.__init__(self,name=self.mgl.fpath.split(os.sep)[-1])
        pass
    def run(self):
        print('='*5)
        print('starting:%s'%threading.current_thread().name)
        # print('reading:',self.mgl.fpath)
        # get_record(self.accid,self.mgl,self.savedir)
        get_record(self.accid_li,self.mgl,self.savedir)
        print('-'*5)
        pass
if __name__=='__main__':
    for j in fileli:
        th=GetGLThread(accid_li,MGL(j,'Sheet1',0,glid_index=['账簿编码','期间年','期间月','凭证编码']))
        th.start()
    pass

#!/usr/bin/env python
# coding=utf-8
'''
Acct是对Account的抽象,会计账户;
Gele是对General Ledger的抽象,序时账.
'''
from autk.financialtk.modules.mortalGL import MGL
class Acct:
    '''
    Acct is short for Account, with two main attributes, 'name' and 'accid'.
    '''
    def __init__(self,accid=r'6001',name='主营业务收入'):
        '''
        class of Account.
        初始化，传入科目名称，科目编码，增加方向，账户类别。
        '''
        self.name=name # 科目名称
        try:
            self.accid=str(int(accid)) # 科目编码
        except:
            self.accid=str(accid)
        # self.isdr=True # 借方主导的科目,共同类科目也归入此类。
        # self.iscr=-self.isdr # 贷方主导的科目
        # self.cata='' # account catagory
class Gele(MGL): # GeneralLedger
    '''
    General Ledgers in a sheet from an Excel Workbook.
    Default sheet name is '表页-1'.
    从审计系统里,读取账套数据，查询凭证，导出序时账和余额表Excel之后，进行抽象，得到zhgl类。
    Gele.sample()方法是有缺陷的,对某方向上累计金额未0的科目无效,会报错,原因在于acct_sum的计算会遇到0/0的情况.
    Gele.sample()的bug已经修复,但有个不足之处,从算法上讲,该抽样的"目标累计金额"是根据序时账GL的发生额计算出的,其实由于账务系统可能存在同账户金额互相转的情况,可能存在GL发生额大于实际发生额的情况,解决之道是从余额表TB获取计算"目标累计金额",如此可获得准确的该账户借方/贷方的发生额.
    '''
    def __init__(self,fpath='',shtna=r'表页-1',title=3,glid_index=[],auto=False):
        import os
        self.fpath=fpath
        self.sheetname=shtna
        self.shtna=shtna
        self.title=title
        self.glname=str(self.fpath.split(os.sep)[-1])
        self.glid_index=glid_index
        self.data=None
        self.sample_data=None
        # from pandas import read_excel
        # self.cols=read_excel(self.fpath,sheet_name=self.sheetname,header=self.title,engine='openpyxl').columns
        # self.cols=['凭证日期', '字', '号', '摘要', '科目编号', '科目全路径', '借方发生金额', '贷方发生金额', '汇率', '外币金额', '外币名称', '数量额', '单价', '计量单位', '核算编号', '核算名称']
        # glid 是GL的主键。
        # self.colsk=['凭证日期', '字', '号', '摘要', 'glid', '科目编号', '科目全路径', '借方发生金额', '贷方发生金额', '汇率', '外币金额', '外币名称', '数量额', '单价', '计量单位', '核算编号', '核算名称']
        print('Before GL initialized，column "glid" should be added.')
        print('=====\nGL name:%s'%self.glname)
        print('GL path:\n',self.fpath)
        print('GL sheet:\t',self.sheetname)
        # print('GL columns:\n',self.cols)
        return
    def getshtli(self): 
        from openpyxl import load_workbook
        return load_workbook(self.fpath).sheetnames
    def getcols(self):
        from pandas import read_excel
        col=read_excel(self.fpath,sheet_name=self.sheetname,header=self.title,engine='openpyxl').columns
        return list(col)
    def getdata(self,fillna=False): # get raw data without column "glid".
        from pandas import read_excel
        d1=read_excel(self.fpath,sheet_name=self.sheetname,header=self.title,engine='openpyxl')
        if fillna==True:
            d1=d1.fillna(float(0.0))
        else:
            pass
        return d1
    def iterate_entry_record(self):
        '''
        Iterate record of raw data of General Ledgers and add the column 'glid'.
        Return a generator of EntryRecord, of which pure GL(with new column 'glid') consists.
        This method requires 'self.getdata(self,fillna=True)' and is required by 'self.getgldata(self)'.
        '''
        from autk.financialtk.journal import EntryRecord
        data=self.getdata(fillna=True)
        def start_record_iteration(indf):
            for i in indf.iterrows():
                yield i
        def get_entry_records(df_iterrows_element):
            for i in df_iterrows_element:
                entry_record=EntryRecord(i)
                if entry_record.__dict__ != {}:
                    yield entry_record
                else:
                    pass
        return get_entry_records(start_record_iteration(data))
    def getgldata(self):
        '''
        Get General Ledger Data with column 'glid' added.
        This method requires 'self.iterate_entry_record(self)'.
        '''
        from pandas import DataFrame
        entry_record=self.iterate_entry_record()
        def get_record_detail():
            for i in entry_record:
                yield i.__dict__
        return DataFrame(get_record_detail())
    def get_journal_entries(self):
        '''
        Iterate each filteration of GL (returned by 'self.getgldata(self)') with 'glid'.
        Return a generator of JEntry(Journal Entry).
        This method requires 'self.getgldata(self)' and is (planned to be) required by 'self.getAccGl(self,account,pure=True)'.
        '''
        from autk.financialtk.journal import JEntry,EntryRecord
        data=self.getgldata()
        def iterate_gl_filter(data):
            glid_li=data['glid'].drop_duplicates()
            glid_li=list(glid_li)
            for i in glid_li:
                one_entry_df=data[data['glid']==i]
                yield [i,one_entry_df]
        for j in iterate_gl_filter(data):
            yield JEntry(j[0],j[1])
    # @classmethod
    def zh_filter(self,regitem,label=r'',match=False):
        '''
        Filter the specific column of the table by regular expression.
        '''
        import re
        indf=self.getdata()
        # indf=self.getgldata() # Do not use the return of self.getgldata() as the input of this method.
        reg=re.compile(regitem)
        fli=[]
        for i in list(indf[label].drop_duplicates()):
            if match==False:
                b=re.search(reg,str(i))
            else:
                b=re.match(reg,str(i))
            if b != None:
                fli.append(i)
            else:
                pass
        from pandas import DataFrame,concat
        ftableli=[]
        for j in fli:
            ftable_fake=indf[indf[label]==j]
            ftableli.append(ftable_fake)
            continue
        if len(ftableli)==0:
            ftable=DataFrame([],index=[],columns=self.getcol())
            pass
        else:
            ftable=concat(ftableli,axis=0,join='outer')
        return ftable
    # def refilter(self,indf,regitem,label=r'',match=False):
    #     '''
    #     filter the specific column of a input table by regular expression.
    #     '''
    #     import re
    #     reg=re.compile(regitem)
    #     fli=[]
    #     for i in list(indf[label].drop_duplicates()):
    #         if match==False:
    #             b=re.search(reg,str(i))
    #         else:
    #             b=re.match(reg,str(i))
    #         if b != None:
    #             fli.append(i)
    #         else:
    #             pass
    #     from pandas import DataFrame,concat
    #     tableli=[]
    #     for j in fli:
    #         ftable_fake=indf[indf[label]==j]
    #         tableli.append(ftable_fake)
    #         continue
    #     ftable=concat(tableli,axis=0,join='outer')
    #     return ftable
    def salaryana(self):
        '''
        职工薪酬计提和分配的核查.
        '''
        salary_id=r'2211'
        expense_id=['6601','6602','5001','5301','2211','1604']
        def transReg(accid):
            return str(r'^')+str(accid)+str(r'.*')
        s1=self.zh_filter(transReg(salary_id),label=r'科目编号')
        # s1=self.refilter(s1,'0','借方发生金额')
        s1=s1[s1['借方发生金额']==0.0]
        eli=[]
        for i in expense_id:
            print('-·'*5)
            print('filtering %s'%i)
            e1=self.zh_filter(transReg(i),label=r'科目编号')
            if e1.shape[0]==0:
                print('no results after filter: %s'%i)
                pass
            else:
                # print(e1.columns)
                # e1=self.refilter(e1,'0','贷方发生金额')
                e1=e1[e1['贷方发生金额']==0.0]
                eli.append(e1)
            continue
        from pandas import concat,pivot_table
        e_concat=concat(eli,axis=0,join='outer')
        salaryTable=concat([s1,e_concat],axis=0,join='outer')
        # sumtable=pivot_table(salaryTable,values=['借方发生金额','贷方发生金额'],index='科目编号',columns=['借方发生金额'],aggfunc='sum')
        sumtable1=pivot_table(salaryTable,values=['借方发生金额'],index='科目编号',columns=[],aggfunc='sum')
        print(sumtable1.sum(axis=0))
        sumtable2=pivot_table(salaryTable,values=['贷方发生金额'],index='科目编号',columns=[],aggfunc='sum')
        print(sumtable2.sum(axis=0))
        print(sumtable2.sum(axis=0)-sumtable1.sum(axis=0))
        for i in list(salaryTable['GLID'].drop_duplicates()):
            dr_sum=(pivot_table(salaryTable[salaryTable['GLID']==i],values=['借方发生金额'],index=['GLID'],columns=[],aggfunc='sum'))
            cr_sum=(pivot_table(salaryTable[salaryTable['GLID']==i],values=['贷方发生金额'],index=['GLID'],columns=[],aggfunc='sum'))
            # print(dr_sum)
            # print(cr_sum)
            drcrdif=(dr_sum.sum(axis=0)['借方发生金额']-cr_sum.sum(axis=0)['贷方发生金额'])
            if drcrdif != 0:
                # print('='*5)
                print(i)
                # print(dr_sum)
                # print(cr_sum)
        return 
    def getAccGl(self,account,pure=True):
        '''
        account is an instance object of class Acct.
        得到某科目的所有记录,默认不含对方科目.
        '''
        regitem=r'^'+str(account.accid)+r'.*'
        print(regitem)
        km=self.zh_filter(regitem,r'科目编号') # km is the Chinese Phonetical Alphabets "Ke Mu".
        idli=list(km['glid'].drop_duplicates())
        def itersearch(idli):
            for i in idli:
                yield self.zh_filter(i,'glid')
                continue
            pass
        if pure == True:
            from pandas import concat
            resu=concat(itersearch(idli),axis=0,join='outer')
            pass
        else:
            resu=km
        self.data=resu
        return resu
    def accum_sample(self,acct_id,filterIdCol,acquired_rate=0.81,drcrdesc=[r'借方发生金额',r'贷方发生金额']):
        '''
        'acct_id',short for 'account id number', can be regular expression to filter in the column of 'filterIdCol'.
        'acquired_rate' is the accumulate sum rate that is required by the manager.
        'dr' is short for Debit while 'cr' for credit.
        Finally return a pandas.DataFrame as a sample.
        '''
        from pandas import read_excel
        # gl=read_excel(self.fpath,sheet_name=self.sheetname)
        gl=self.getdata()
        # regitem=r'^'+str(acct_id)+r'.*'
        regitem=str(acct_id)
        theAcct=gl.filter(regitem,filterIdCol) # 筛出被抽样的科目.
        print('this account shape:',theAcct.shape)
        # from autk.zhchart import ChartAccount
        # chac=ChartAccount()
        # acctli=chac.getAcct()
        from pandas import DataFrame
        # acct_sum=DataFrame([acct_li[2],acct_li[3]],index=inAcct.accid,columns=drcrdesc)
        acct_sum=theAcct[drcrdesc].sum(axis=0) # 被指定的科目借方贷方求和.这里要修改,要从余额表读取此数.
        sub_count=theAcct[drcrdesc].count(axis=0)
        averAmount=acct_sum/sub_count # 平均每个样本的金额
        target_sum=acct_sum*acquired_rate # 依据目标金额比例,确定目标累计合计金额.
        # start_nums=target_sum/averAmount # 计算初始样本容量,如acct_sum为0,start_nums会报错.
        print(target_sum)
        print(type(target_sum))
        def get_dr_sample():
            # 开始抽借方样本:
            n_dr_start=int(target_sum[0]/averAmount[0]/2)
            # n_dr_start=1
            dr_sample=theAcct.nlargest(n=n_dr_start,columns=[drcrdesc[0]],keep='last')
            # print('dr sample sum:',dr_sample[drcrdesc].sum(axis=0))
            dr_sam_rate=dr_sample[drcrdesc[0]].sum(axis=0)/acct_sum[0]
            while dr_sam_rate<acquired_rate:
                n_dr_start+=1
                dr_sample=theAcct.nlargest(n=n_dr_start,columns=[drcrdesc[0]],keep='last')
                dr_sam_rate=dr_sample[drcrdesc[0]].sum(axis=0)/acct_sum[0]
                continue
            return dr_sample
        def get_cr_sample():
            # 开始抽贷方样本:
            n_cr_start=int(target_sum[1]/averAmount[1]/2)
            # n_cr_start=1
            cr_sample=theAcct.nlargest(n=n_cr_start,columns=[drcrdesc[1]],keep='last')
            cr_sam_rate=cr_sample[drcrdesc[1]].sum(axis=0)/acct_sum[1]
            while cr_sam_rate<acquired_rate:
                n_cr_start+=1
                cr_sample=theAcct.nlargest(n=n_cr_start,columns=[drcrdesc[1]],keep='last')
                cr_sam_rate=cr_sample[drcrdesc[1]].sum(axis=0)/acct_sum[1]
                continue
            return cr_sample
        # 将借方贷方样本拼接在一起:
        if acct_sum[0] !=0: 
            dr=get_dr_sample()
            pass
        else:
            dr=DataFrame([],columns=theAcct.columns)
            pass
        if acct_sum[1] !=0:
            cr=get_cr_sample()
            pass
        else:
            cr=DataFrame([],columns=theAcct.columns)
            pass
        from pandas import concat
        final_sample=concat([dr,cr],axis=0,join='outer',ignore_index=True)
        final_sample=final_sample.reset_index(drop=True)
        self.sample_data=final_sample
        return final_sample
    def write_sample(self,filterIdCol,account,savedir):
        '''
        account is an instance object of class Acct with attributes of 'accid' and 'name'.
        '''
        from openpyxl import load_workbook,Workbook
        wb=Workbook()
        wb.save(savedir)
        wb.close()
        wb=load_workbook(savedir)
        from pandas import ExcelWriter
        wter=ExcelWriter(savedir,engine='openpyxl')
        wter.book=wb
        s1=self.accum_sample(account.accid,filterIdCol,acquired_rate=0.81,drcrdesc=[r'借方发生金额',r'贷方发生金额'])
        s1.to_excel(wter,sheet_name=account.name)
        wter.save()
        wter.close()

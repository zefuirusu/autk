#!/usr/bin/env python
# coding=utf-8
'''
Plan before sampling.
Materiality
tolerance
'''
class Materiality:
    '''
    重要性水平
    '''
    def __init__(self,xldir):
        self.xldir=xldir
        self.metadata=self.loadxl()
        pass
    def loadxl(self):
        from pandas import read_excel
        m=read_excel(self.xldir,sheet_name='materiality',header=0,engine='openpyxl')
        m=dict(zip(m.iloc[:,0],m.iloc[:,1]))
        return m

class PreScan:
    def __init__(self,wshtdir):
        '''
        Working Sheet Directory is acquired.
        '''
        self.wshtdir=wshtdir
        return
    def walk(self):
        import os
        import re
        for i,j,k in os.walk(self.wshtdir):
            for na in k:
                regitem=re.compile(r'xls[xm]$')
                if re.search(regitem,na) != None:
                    fpath=os.path.abspath(os.path.join(i,na))
                    # print('-'*5)
                    # print(fpath)
                    # print(na)
                    yield fpath
                    pass
                else:
                    pass
        return
    def scan(self):
        import os
        import re
        from openpyxl import load_workbook
        from pandas import read_excel
        fileli=self.walk()
        for i in fileli:
            wb=load_workbook(i,keep_vba=True)
            shtli=wb.sheetnames
            # print(shtli)
            print('='*5)
            print(i.split(os.sep)[-1])
            for j in shtli:
                if re.search('审定',j) !=None:
                    print(j)
                    d=read_excel(i,sheet_name=j)
                    # ws=wb.get_sheet_by_name(j)
                    # print(ws)
                    for n in range(0,d.shape[0]):
                        # for m in range(1,10):
                        #     print(ws.cell(n,1).value)
                        print(d.iloc[n,1])
                        continue
                    pass
                else:
                    # print('failed!')
                    pass
                print('%s is finished.'%j)
            print('-'*5)
            continue
        return
#

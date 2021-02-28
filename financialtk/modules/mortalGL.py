#!/usr/bin/env python
# coding=utf-8
'''
EntryRecord: EntryRecord;
JEntry: Journal Entry;
MGL: Mortal General Ledger.
'''
import re
from numpy import nan
from pandas import read_excel,DataFrame,concat
def transType(element):
    '''
    transfer an float/integer object into string.
    '''
    if element is nan:
    # if isinstance(element,type(nan)):
        print('woc!!! %s is nan!'%element )
        element=str(int(0))
        print("It's ok, nan has been transformed into string.")
    else:
        if isinstance(element,float):
            element=str(int(element))
        elif isinstance(element,int):
            element=str(element)
        elif isinstance(element,str):
            element=element
    return element
class EntryRecord:
    def __init__(self,df_iterrows_element,glid_cols=[0,1,2]):
        self.index=list(df_iterrows_element[1].index)
        def re_match(in_string,what):
            import re
            what=transType(what)
            regex_item=re.compile(in_string)
            return re.match(regex_item,what)
        def get_id_list():
            id_list=[]
            for i in df_iterrows_element[1].index:
                series_element=df_iterrows_element[1][i]
                series_element_index=transType(i)
                if re_match('.*日期.*',series_element_index) != None:
                    id_list.append(series_element)
                elif re_match('.*字.*',series_element_index) != None: 
                    id_list.append(series_element) 
                elif re_match('.*号.*',series_element_index) != None:
                    id_list.append(series_element)
                continue
            return id_list
        def start_init():
            for i in df_iterrows_element[1].index:
                series_element=df_iterrows_element[1][i]
                series_element_index=transType(i)
                if re_match('.*借方?.*',series_element_index) != None:
                    self.dr_amount=series_element
                elif re_match('.*贷方?.*',series_element_index) != None:
                    self.cr_amount=series_element
                elif re_match(r'.*科目编[号,码].*',series_element_index) != None:
                    self.accid=transType(series_element)
                    self.top_accid=self.accid[0:4]
                elif re_match(r'.*科目(.*路径.*|.*名称.*).*',series_element_index) != None:
                    self.accna=series_element
                elif re_match('.*对[方,应]科目.*',series_element_index) != None:
                    self.opposite=series_element
                elif re_match('.*日期.*',series_element_index) != None:
                    self.date=series_element
                elif re_match('.*月.*',series_element_index) != None:
                    self.month=series_element
                elif re_match(r'.*(摘要|文本).*',series_element_index) != None:
                    self.scan=series_element
                else:
                    pass
                continue
            pass
        id_list=get_id_list()
        if 'glid' in self.index:
            self.glid=df_iterrows_element[1]['glid']
            start_init()
        else:
            if nan not in id_list:
                self.glid=r'-'.join(map(transType,id_list))
                start_init()
            else:
                print('entry record initialize failed.')
        pass
    def getdata(self):
        from pandas import DataFrame
        return [self.glid,DataFrame(self.__dict__)]
    pass
# class JsonEntryRecord(EntryRecord):
#     def __init__(self,df_iterrows_element,glid_cols=[0,1,2]):
#         pass
class JEntry:
    '''
    Journal Entry.
    '''
    # __slots__=('glid','date','scan','debit','credit') # static limit of S ?
    def __init__(self,glid,one_entry_df):
        dr_sum=one_entry_df['dr_amount'].sum(axis=0)
        dr_sum=float(dr_sum)
        cr_sum=one_entry_df['cr_amount'].sum(axis=0)
        cr_sum=float(cr_sum)
        if dr_sum-cr_sum>=-0.009 and dr_sum-cr_sum<=0.009:
            pass
            # print('OK! Debit = Credit!')
        else:
            print('%s ! Dr/Cr not balenced!'%glid)
            # print('dr/cr:',dr_sum,'/',cr_sum)
            pass
        self.glid=glid # glid is the unique id of JEntry.
        self.date=self.glid[0:10] # date of the Journal Entry Recording.
        scan=one_entry_df['scan']
        if len(scan.drop_duplicates()) != 1:
            self.scan=list(scan)
        else:
            self.scan=list(scan.drop_duplicates())[0]# summary or comments of each EntryRecord.

        self.debit=[]
        self.credit=[]
        for i in one_entry_df.iterrows():
            er=i[1]
            if er.cr_amount==0.00:
                self.debit.append({"accid":er.accid,"amount":er.dr_amount,"accna":er.accna})
            elif er.dr_amount==0.00: # er.dr_amount==0:
                self.credit.append({"accid":er.accid,"amount":er.cr_amount,"accna":er.accna})
            pass
        pass
class MGL:
    '''
    Mortal General Ledger, a template class of General Ledger.
    '''
    def __init__(self,fpath='',shtna='Sheet1',title=0,glid_index=[],auto=False):
        '''
        Three key elements of an General Ledger File is the path, sheet name and the column title location.
        '''
        self.fpath=fpath
        self.shtna=shtna
        self.title=title
        self.glid_index=glid_index
        self.data=None
        self.sample_data=None
        self.entry_info=None
        self.gl_matrix=None
        if self.data is not None:
            if 'glid' in self.data.columns:
                self.glid_list=list(self.data['glid'].drop_duplicates())
                pass
            else:
                self.glid_list=[]
                pass
            pass
        else:
            self.glid_list=None
        pass
        if auto == True:
            self.load_raw_data()
            if self.glid_index == []:
                return
            else:
                self.set_glid(self.glid_index)
                pass
            pass
        else:
            print('You need load data.')
            pass
    def load_df(self,in_df):
        '''
        Load outsource DataFrame data and return Nothing.
        '''
        self.data=in_df
        # return self.data
        pass
    def load_raw_data(self):
        '''
        Load original DataFrame data and return Nothing.
        '''
        self.data=read_excel(self.fpath,sheet_name=self.shtna,header=self.title,engine='openpyxl')
        if 'glid' in self.data.columns :
            self.glid_list=list(self.data['glid'].drop_duplicates())
        pass
    def getshtli(self): 
        from openpyxl import load_workbook
        return load_workbook(self.fpath).sheetnames
    def get_cols(self):
        if self.data is not None:
            return list(self.data.columns)
        else:
            # self.load_raw_data()
            # return list(self.data.columns)
            cols=list(self.get_raw_data().columns)
            print('data has not been loaded. columns in worksheet are:')
            print(cols)
            return cols
    def get_gl_matrix(self,over_write=False,drcr=['借方本币','贷方本币']):
        '''
        Get a matrix of General Ledger.
        '''
        if self.data is not None:
            if 'glid' in self.data.columns:
                self.data['drcr_value']=self.data[drcr[0]]-self.data[drcr[1]]
                gl_matrix=self.data.pivot_table(values=['drcr_value'],index=['glid'],columns=['科目编码'])
                if over_write==True:
                    self.gl_matrix=gl_matrix
                else:
                    pass
                return gl_matrix
                pass
            else:
                print('Woc! glid is not set.')
                if self.glid_index != []:
                    self.set_glid(self.glid_index)
                    pass
                else:
                    print('Set glid_index first!')
                    return
                pass
        else:
            print('Woc! raw data is not loaded. Load it and set glid first.')
            # self.load_raw_data()
        pass
    def get_raw_data(self):
        '''
        Get and return original DataFrame data.
        This method does not LOAD original DataFrame data.
        '''
        from pandas import read_excel
        data=read_excel(self.fpath,sheet_name=self.shtna,header=self.title,engine='openpyxl')
        return data
        pass
    def set_glid(self,glid_index=[]):
        '''
        parameters:
            glid_index, values, not numbers.
        '''
        if glid_index is []:
            print("Pass glid_index as argument first!")
            glid_index=self.glid_index
        else:
            self.glid_index=glid_index
            print('self.glid_index is set to:',glid_index)
            pass
        print(glid_index)
        if glid_index == []:
            print("Pass glid_index as argument first!")
            return
        else:
            print("OK, glid_index is not empty.")
            print('glid_index:%s'%glid_index)
            pass
        if self.data is None:
            print('You need load data first!')
            self.load_raw_data()
            print('Oh, raw data has been loaded.')
            pass
        else:
            def get_glid_li():
                for i in self.data.iterrows():
                    row=i[1]
                    glid=[]
                    for j in glid_index:
                        a_index=row[j]
                        # a_index=str(a_index)
                        a_index=transType(a_index)
                        glid.append(a_index)
                    glid='-'.join(glid)
                    # i['glid']=glid
                    d=dict(row)
                    d['glid']=glid
                    # yield {glid:row}
                    yield d
            data=DataFrame(get_glid_li())
            self.load_df(data)
            self.glid_list=list(self.data['glid'].drop_duplicates())
            print('glid is set.',self.data.shape)
            return data
            pass
        pass
    def iterate_jsfile(self,jspath):
        import re
        import json
        # from pandas import Series,concat
        with open(jspath,mode='r',encoding='utf-8') as f:
            jslines=f.readlines()
            pass
        for i in jslines:
            js_record=re.sub(re.compile(r'\s*$'),'',i)
            yield json.loads(js_record,encoding='utf-8')
        pass
    def iterate_record(self):
        '''
        iterate rows of self.data(DataFrame) and yield data of EntryRecord object.
        '''
        if self.data is not None:
            for i in self.data.iterrows():
                yield EntryRecord(i)
                continue
            pass
        else:
            self.load_raw_data()
            for i in self.data.iterrows():
                yield EntryRecord(i)
                continue
            pass
        pass
    def iterate_json_record(self):
        '''
        iterate rows of self.data(DataFrame) and yield data of json(python dict) of format.
        '''
        # self.load_raw_data()
        if self.data is not None:
            for i in self.data.iterrows():
                yield dict(i[1])
                continue
            pass
        else:
            self.load_raw_data()
            for i in self.data.iterrows():
                yield dict(i[1])
                continue
            pass
    def iterate_filter(self,regitem,label=r'',match=False):
        '''
        iterate row_data of self.data and yield.
        yield row_data as Series.
        '''
        import re
        def start_iterate():
            reg=re.compile(regitem)
            indf=self.data
            for i in self.data.iterrows():
                row_index=i[0]
                row_data=i[1]
                key_str=str(row_data[label])
                if match==False:
                    b=re.search(reg,key_str)
                else:
                    b=re.match(reg,key_str)
                if b is not None:
                    # glid=r'-'.join([row_data['账簿编码'],row_data['期间年'],row_data['期间月'],row_data['凭证编码']])
                    # yield {glid,row_data}
                    yield row_data
                    pass
                else:
                    # yield None
                    pass
                continue
            pass
        if self.data is not None:
            return start_iterate()
        else:
            self.load_raw_data()
            return start_iterate()
            pass
    def filter(self,regitem,label=r'',match=False,over_write=False):
        '''
        Filter the specific column of the table by regular expression.
        '''
        if self.data is None:
            print('Load data first!')
            pass
        else:
            print('Start filtering......')
        # self.load_raw_data()
        # self.set_glid()
        # import re
        indf=self.data
        reg=re.compile(regitem)
        # fli=[]
        def iterate_labels():
            '''
            Iterate and filter labels.
            '''
            for i in list(indf[label].drop_duplicates()):
                if match==False:
                    b=re.search(reg,str(i))
                else:
                    b=re.match(reg,str(i))
                if b is not None:
                    # fli.append(i)
                    yield i
                else:
                    pass
        from pandas import DataFrame,concat
        # ftableli=[]
        # for j in fli:
        def get_filter_tables():
            '''
            Filter and yield self.data by labels got so as to concatenate.
            '''
            for j in iterate_labels():
                ftable_fake=indf[indf[label]==j]
                # ftableli.append(ftable_fake)
                yield ftable_fake
                continue
        if len(list(get_filter_tables()))==0:
        # if len(ftableli)==0:
            ftable=DataFrame([],index=[],columns=self.get_cols())
            pass
        else:
            # ftable=concat(ftableli,axis=0,join='outer')
            ftable=concat(get_filter_tables(),axis=0,join='outer')
            pass
        if over_write == False:
            return ftable
        else:
            self.load_df(ftable)
            return self.data
    def getAcct(self,accid_item,accid_label='科目编码',over_write=False,pure=False):
        '''
        Get all and full records about given 'accid'.
        glid must be set first.
        '''
        if 'glid' in self.get_cols():
            pass
        else:
            self.set_glid(self.glid_index)
        def get_relevant_rows(id_li):
            # self.set_glid()
            for i in id_li:
                yield self.data[self.data['glid']==i]
        if pure==True:
            acct_data=self.filter(accid_item,accid_label,match=False,over_write=False)
            pass
        else:
            # self.set_glid()
            id_li=list(self.filter(accid_item,accid_label,match=False,over_write=False)['glid'].drop_duplicates())
            acct_data=concat(get_relevant_rows(id_li),axis=0,join='outer')
            pass
        if over_write == False:
            return acct_data
        else:
            self.load_df(acct_data)
            return self.data
    def rand_sample(self,ss=None, percent=None, replace=False, weights=None, random_state=None, axis=None,over_write=False):
        '''
        DataFrame.sample(n=None, frac=None, replace=False, weights=None, random_state=None, axis=None)
        parameters:
            ss:sample_size;
            percent:percentage of total to sample;
            axis: 0 for rows_sampling and 1 for colums_sampling;
            over_write: if True, self.sample_data will be over_write and therefore replaced. Currently, this parameter does nothing, once rand_sample(), self.sample_data will be over written.
        '''
        if self.data is not None:
            self.sample_data=self.data.sample(n=ss, frac=percent, replace=False, weights=None, random_state=None, axis=None)
            return self.sample_data
        else:
            self.load_raw_data()
            self.sample_data=self.data.sample(n=ss, frac=percent, replace=False, weights=None, random_state=None, axis=None)
            return self.sample_data
        pass
    def woc(self):
        if self.data is None:
            print('Load data first!')
            return
        else:
            pass

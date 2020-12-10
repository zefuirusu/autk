#!/usr/bin/env python
# coding=utf-8
'''
Account Stated 明细账
'''
class ActStated:
    def __init__(self,fdir,shtna=r'表页-1',title=3):
        from os import sep
        self.path=fdir
        self.filename=str(self.path.split(sep)[-1])
        self.sheetname=shtna
        self.title=title
        self.cols=['抽', '凭证日期', '字', '号', '摘要', '对方科目', '借方发生额', '贷方发生额', '方向', '余额']
        self.colsk=['抽', '凭证日期', '字', '号', 'glid', '摘要', '对方科目', '借方发生额', '贷方发生额', '方向', '余额']
        self.acctcol=['借方发生额', '贷方发生额']
        return
    def getshtli(self):
        from openpyxl import load_workbook
        return load_workbook(self.path).sheetnames
    def getcol(self):
        from pandas import read_excel
        col=read_excel(self.path,sheet_name=self.sheetname,header=self.title).columns
        return list(col)
    def getAcctData(self):
        '''
        [start_amount,debit_amount,credit_amount,end_amount]
        '''
        rawdata=self.getdata(rawdata=True)
        puredata=self.getdata()
        start_amtype=list(rawdata['方向'])[0]
        end_amtype=list(rawdata['方向'])[-1]
        start_amount=rawdata.iat[0,rawdata.shape[1]-1]
        end_amount=rawdata.iat[rawdata.shape[0]-1,rawdata.shape[1]-1]
        debit_amount=puredata[self.acctcol[0]].sum(axis=0)
        credit_amount=puredata[self.acctcol[1]].sum(axis=0)
        return [[start_amount,debit_amount,credit_amount,end_amount],[start_amtype,end_amtype]]
    def getdata(self,rawdata=False):
        from pandas import read_excel
        d=read_excel(self.path,sheet_name=self.sheetname,header=self.title)
        if rawdata==True:
            return d
        else:
            newIndex=d['字'].dropna().index
            d=d.iloc[newIndex,:].reset_index(drop=True)
            return d
    def randSample(self,acquired_rate=0.81,drcrdesc=['借方发生额', '贷方发生额']):
        data=self.getdata()
        s=data.sample(frac=acquired_rate,axis=0,random_state=None,replace=False)
        return s
    def simpSample(self):
        from pandas import concat
        data=self.getdata()
        dr=data.nlargest(n=10,columns=self.acctcol[0],keep='last')
        cr=data.nlargest(n=10,columns=self.acctcol[1],keep='last')
        s=concat([dr,cr],axis=0,join='inner')
        if str('glid') in s.columns:
            glid_li=s['glid'].drop_duplicates().sort_values(ascending=True) # ascending=True为升序排列.
            for i in glid_li:
                print(i)
                continue
            pass
        print('='*5)
        print('Simple Sample Summary for %s'%self.filename)
        print('Sample Volume:')
        print(s[self.acctcol].sum(axis=0))
        return s
